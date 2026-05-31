from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from ocr_price.inventory import (
    InventoryItem,
    apply_inventory_to_project,
    build_inventory_review,
    inventory_items_from_review,
    load_inventory_from_sources,
    parse_inventory_text,
)


def _build_quote_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "报价表"

    # Mill header used by inventory writeback.
    ws.cell(row=1, column=5, value="目标钢厂")

    # Row 12 is treated as rebar by current row mapping logic.
    ws.cell(row=12, column=2, value="12")
    ws.cell(row=12, column=3, value="9")
    ws.cell(row=12, column=4, value="HRB400E")

    # Another row to verify clear_existing_colors removes stale fill.
    ws.cell(row=13, column=2, value="14")
    ws.cell(row=13, column=3, value="9")
    ws.cell(row=13, column=4, value="HRB400E")
    ws.cell(row=13, column=5).fill = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    )

    wb.save(path)


def test_apply_inventory_uses_confirmed_mapping_and_clears_existing_colors(tmp_path: Path):
    project_path = tmp_path / "project.xlsx"
    _build_quote_workbook(project_path)

    mapping_path = tmp_path / "mapping.json"
    mapping_rows = [
        {
            "项目文件Sheet": "目标钢厂",
            "最新清单厂家Sheet": "来源钢厂",
            "状态": "已确认匹配",
            "说明": "",
        }
    ]
    mapping_path.write_text(json.dumps(mapping_rows, ensure_ascii=False), encoding="utf-8")

    mill_inventories = {
        "来源钢厂": [
            InventoryItem(
                product="螺纹",
                spec="12",
                length="9",
                material="HRB400E",
                status="告警",
            )
        ]
    }

    result = apply_inventory_to_project(
        project_excel=project_path,
        mill_inventories=mill_inventories,
        sheet_name="报价表",
        mapping_json_path=mapping_path,
        clear_existing_colors=True,
    )

    assert result["status"] == "ok"
    assert result["applied_count"] == 1
    assert result["cleared_count"] >= 2

    applied = result["applied"][0]
    assert applied["mill"] == "来源钢厂"
    assert applied["sheet_mill"] == "目标钢厂"
    assert applied["product"] == "螺纹"
    assert applied["spec"] == "12"
    assert applied["length"] == "9"
    assert applied["material"] == "HRB400E"
    assert applied["status"] == "告警"
    assert applied["cell"] == "E12"

    wb = load_workbook(project_path)
    ws = wb["报价表"]
    updated_cell = ws.cell(row=12, column=5)
    assert updated_cell.fill.fill_type == "solid"
    assert updated_cell.fill.start_color.rgb in {"00FFC000", "FFC000"}

    cleared_cell = ws.cell(row=13, column=5)
    assert cleared_cell.fill.fill_type is None


def test_load_inventory_from_vision_json_preserves_structured_statuses(tmp_path: Path):
    source = tmp_path / "ocr价格提取_徐钢.json"
    source.write_text(
        json.dumps(
            {
                "meta": {"input_file": str(tmp_path / "徐钢.jpg")},
                "_vision_result": {
                    "库存情况": [
                        {"规格": "9米螺纹14E", "状态": "告警", "原始描述": "极少"},
                        {"规格": "12米螺纹16E", "状态": "告警", "原始描述": "28件"},
                        {"规格": "盘螺8E", "状态": "告警", "原始描述": "16件"},
                    ]
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    items = load_inventory_from_sources([source], "徐钢")

    assert InventoryItem(
        product="螺纹",
        spec="14",
        length="9",
        material=None,
        status="告警",
        note="极少",
    ) in items
    assert InventoryItem(
        product="螺纹",
        spec="16",
        length="12",
        material=None,
        status="告警",
        note="28件",
    ) in items
    assert InventoryItem(
        product="盘螺",
        spec="8",
        length=None,
        material=None,
        status="告警",
        note="16件",
    ) in items


def test_load_inventory_deduplicates_structured_items_by_semantic_key(tmp_path: Path):
    source = tmp_path / "ocr价格提取_徐钢.json"
    source.write_text(
        json.dumps(
            {
                "meta": {"input_file": str(tmp_path / "徐钢.jpg")},
                "_vision_result": {
                    "库存情况": [
                        {"规格": "9米螺纹12E", "状态": "充足", "原始描述": ""},
                        {"规格": "9米螺纹12E", "状态": "告警", "原始描述": "极少"},
                        {"规格": "蚌埠螺纹9m 14E", "状态": "充足", "原始描述": ""},
                        {"规格": "9米螺纹14E (蚌埠)", "状态": "告警", "原始描述": "14E配"},
                    ]
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    items = load_inventory_from_sources([source], "徐钢")

    matching_12 = [
        item for item in items
        if item.product == "螺纹" and item.spec == "12" and item.length == "9"
    ]
    matching_14 = [
        item for item in items
        if item.product == "螺纹"
        and item.spec == "14"
        and item.length == "9"
        and item.warehouse == "蚌埠"
    ]
    assert len(matching_12) == 1
    assert matching_12[0].status == "充足"
    assert len(matching_14) == 1
    assert matching_14[0].status == "充足"


def test_build_inventory_review_reports_duplicates_conflicts_and_selection(
    tmp_path: Path,
):
    source = tmp_path / "ocr价格提取_徐钢.json"
    source.write_text(
        json.dumps(
            {
                "meta": {"input_file": str(tmp_path / "徐钢.jpg")},
                "_vision_result": {
                    "库存情况": [
                        {"规格": "9米螺纹12E", "状态": "充足", "原始描述": ""},
                        {"规格": "9米螺纹12E", "状态": "告警", "原始描述": "极少"},
                        {"规格": "盘螺8E", "状态": "告警", "原始描述": "16件"},
                    ]
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    review = build_inventory_review([source])

    assert review["status"] == "ok"
    assert review["raw_count"] == 3
    assert review["selected_count"] == 2
    assert review["duplicate_group_count"] == 1
    assert review["conflict_group_count"] == 1
    conflict = review["conflict_groups"][0]
    assert conflict["company"] == "徐钢"
    assert conflict["product"] == "螺纹"
    assert conflict["spec"] == "12"
    assert conflict["length"] == "9"
    assert conflict["statuses"] == ["充足", "告警"]
    assert conflict["selected"]["status"] == "充足"


def test_build_inventory_review_prefers_original_text_over_stale_top_level_inventory(
    tmp_path: Path,
):
    text = tmp_path / "闽源.txt"
    text.write_text("9米HRB400E规格有：20（配货）", encoding="utf-8")
    source = tmp_path / "ocr价格提取_闽源.json"
    source.write_text(
        json.dumps(
            {
                "meta": {"input_file": str(text)},
                "company": "闽源集团",
                "inventory": [
                    {"规格": "9米螺纹20", "状态": "充足", "原始描述": "20"}
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    review = build_inventory_review([source])

    assert review["selected"] == [
        {
            "company": "闽源集团",
            "source_file": "ocr价格提取_闽源.json",
            "source_spec": "9米HRB400E螺纹20",
            "source_kind": "original_text",
            "confidence_basis": "原始txt解析",
            "source_priority": 100,
            "warehouse": "",
            "product": "螺纹",
            "spec": "20",
            "length": "9",
            "material": "HRB400E",
            "status": "告警",
            "note": "配货",
        }
    ]


def test_inventory_review_selection_uses_confidence_priority_across_sources(
    tmp_path: Path,
):
    stale = tmp_path / "ocr价格提取_闽源_旧.json"
    stale.write_text(
        json.dumps(
            {
                "company": "闽源集团",
                "inventory": [
                    {"规格": "9米HRB400E螺纹20", "状态": "充足", "原始描述": "旧缓存"}
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    text = tmp_path / "闽源.txt"
    text.write_text("9米HRB400E规格有：20（配货）", encoding="utf-8")
    fresh = tmp_path / "ocr价格提取_闽源.json"
    fresh.write_text(
        json.dumps(
            {
                "company": "闽源集团",
                "meta": {"input_file": str(text)},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    review = build_inventory_review([stale, fresh])

    assert review["raw_count"] == 2
    assert review["selected_count"] == 1
    assert review["conflict_group_count"] == 1
    selected = review["selected"][0]
    assert selected["status"] == "告警"
    assert selected["source_kind"] == "original_text"
    assert selected["confidence_basis"] == "原始txt解析"
    assert review["conflict_groups"][0]["selected"]["status"] == "告警"


def test_inventory_items_from_review_keeps_selected_source_evidence(
    tmp_path: Path,
):
    text = tmp_path / "闽源.txt"
    text.write_text("9米HRB400E规格有：20（配货）", encoding="utf-8")
    source = tmp_path / "ocr价格提取_闽源.json"
    source.write_text(
        json.dumps(
            {
                "company": "闽源集团",
                "meta": {"input_file": str(text)},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    review = build_inventory_review([source])

    mill_inventories = inventory_items_from_review(review)

    item = mill_inventories["闽源集团"][0]
    assert item.status == "告警"
    assert item.material == "HRB400E"
    assert item.source_file == "ocr价格提取_闽源.json"
    assert item.source_spec == "9米HRB400E螺纹20"
    assert item.confidence_basis == "原始txt解析"


def test_parse_inventory_text_detects_peihuo_and_chinese_parentheses_material():
    items = parse_inventory_text(
        "9米HRB400E规格有：12、20（配货）\n"
        "12米HRB500E规格有：20（1件）"
    )

    assert InventoryItem(
        product="螺纹",
        spec="20",
        length="9",
        material="HRB400E",
        status="告警",
        note="配货",
    ) in items
    assert InventoryItem(
        product="螺纹",
        spec="20",
        length="12",
        material="HRB500E",
        status="告警",
        note="1件",
    ) in items


def test_apply_inventory_matches_rebar_and_coil_rows(tmp_path: Path):
    project_path = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "报价表"
    ws.cell(row=1, column=5, value="徐钢")

    ws.cell(row=13, column=2, value="8")
    ws.cell(row=13, column=3, value=None)
    ws.cell(row=13, column=4, value="HRB400E")

    ws.cell(row=19, column=2, value="14")
    ws.cell(row=19, column=3, value="9")
    ws.cell(row=19, column=4, value="HRB400E")

    wb.save(project_path)

    result = apply_inventory_to_project(
        project_excel=project_path,
        mill_inventories={
            "徐钢": [
                InventoryItem(
                    product="盘螺",
                    spec="8",
                    length=None,
                    material=None,
                    status="告警",
                    note="16件",
                ),
                InventoryItem(
                    product="螺纹",
                    spec="14",
                    length="9",
                    material=None,
                    status="告警",
                    note="极少",
                ),
            ]
        },
        sheet_name="报价表",
    )

    assert result["status"] == "ok"
    assert result["applied_count"] == 2

    wb = load_workbook(project_path)
    ws = wb["报价表"]
    assert ws.cell(row=13, column=5).fill.fill_type == "solid"
    assert ws.cell(row=13, column=5).fill.start_color.rgb in {"00FFC000", "FFC000"}
    assert ws.cell(row=19, column=5).fill.fill_type == "solid"
    assert ws.cell(row=19, column=5).fill.start_color.rgb in {"00FFC000", "FFC000"}


def test_changjiang_uses_dynamic_warehouse_headers_but_only_factory_and_bengbu(
    tmp_path: Path,
):
    project_path = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "报价表"
    ws.cell(row=1, column=5, value="马长江")

    source = wb.create_sheet("长江")
    source.cell(row=8, column=8, value="蚌埠库")
    source.cell(row=8, column=9, value="网差")
    source.cell(row=8, column=10, value="钢厂")
    source.cell(row=8, column=11, value="网差")
    source.cell(row=8, column=12, value="阜阳库")
    source.cell(row=8, column=13, value="网差")

    ws.cell(row=9, column=5, value="=长江!H8")
    ws.cell(row=9, column=7, value="=长江!J8")
    ws.cell(row=9, column=9, value="=长江!L8")
    ws.cell(row=9, column=2, value="规格")
    ws.cell(row=9, column=3, value="长度（米）")
    ws.cell(row=9, column=4, value="材质")
    ws.cell(row=12, column=2, value="12")
    ws.cell(row=12, column=3, value="9")
    ws.cell(row=12, column=4, value="HRB400E")
    ws.cell(row=12, column=9).fill = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    )
    wb.save(project_path)

    result = apply_inventory_to_project(
        project_excel=project_path,
        mill_inventories={
            "长江": [
                InventoryItem(
                    product="螺纹",
                    spec="12",
                    length="9",
                    material=None,
                    warehouse="蚌埠",
                    status="告警",
                ),
                InventoryItem(
                    product="螺纹",
                    spec="12",
                    length="9",
                    material=None,
                    warehouse="厂内",
                    status="缺货",
                ),
                InventoryItem(
                    product="螺纹",
                    spec="12",
                    length="9",
                    material=None,
                    warehouse="阜阳",
                    status="缺货",
                ),
            ]
        },
        sheet_name="报价表",
        clear_existing_colors=True,
    )

    assert result["status"] == "ok"
    assert result["applied_count"] == 2
    assert {item["cell"] for item in result["applied"]} == {"E12", "G12"}

    wb = load_workbook(project_path)
    ws = wb["报价表"]
    assert ws["E12"].fill.fill_type == "solid"
    assert ws["E12"].fill.start_color.rgb in {"00FFC000", "FFC000"}
    assert ws["G12"].fill.fill_type == "solid"
    assert ws["G12"].fill.start_color.rgb in {"00FF0000", "FF0000"}
    assert ws["I12"].fill.fill_type is None


def test_apply_inventory_reports_each_cell_once_when_generic_and_specific_match(
    tmp_path: Path,
):
    project_path = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "报价表"
    ws.cell(row=1, column=5, value="目标钢厂")
    ws.cell(row=12, column=2, value="20")
    ws.cell(row=12, column=3, value="9")
    ws.cell(row=12, column=4, value="HRB400E")
    wb.save(project_path)

    result = apply_inventory_to_project(
        project_excel=project_path,
        mill_inventories={
            "目标钢厂": [
                InventoryItem(
                    product="螺纹",
                    spec="20",
                    length="9",
                    material=None,
                    status="充足",
                ),
                InventoryItem(
                    product="螺纹",
                    spec="20",
                    length=None,
                    material=None,
                    status="告警",
                ),
            ]
        },
        sheet_name="报价表",
    )

    assert result["applied_count"] == 1
    assert result["applied"][0]["cell"] == "E12"
    assert result["applied"][0]["status"] == "充足"
