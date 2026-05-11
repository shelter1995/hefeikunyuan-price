from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ocr_price.writeback_image_doc import (
    PriceDeviationConfig,
    _check_price_deviation,
    apply_writeback,
)


def _make_project(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "报价表"
    mill = wb.create_sheet("测试钢厂")
    mill["G1"] = "网价[2026-05-11]"
    mill["G3"] = 3300
    mill["G4"] = 3100
    mill["H3"] = 3300
    mill["H4"] = 3100
    wb.save(path)


def _write_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def test_price_deviation_allows_value_within_absolute_and_percent_thresholds():
    result = _check_price_deviation(
        offline_price=3600,
        web_price=3300,
        label="盘螺",
        config=PriceDeviationConfig(abs_tolerance=1000, pct_tolerance=0.20),
    )
    assert result is None


def test_price_deviation_blocks_when_absolute_threshold_exceeded():
    result = _check_price_deviation(
        offline_price=4501,
        web_price=3300,
        label="盘螺",
        config=PriceDeviationConfig(abs_tolerance=1000, pct_tolerance=0.20),
    )
    assert result is not None
    assert "盘螺" in result
    assert "差值=1201" in result


def test_price_deviation_blocks_when_percent_threshold_exceeded():
    result = _check_price_deviation(
        offline_price=4050,
        web_price=3000,
        label="螺纹",
        config=PriceDeviationConfig(abs_tolerance=1000, pct_tolerance=0.20),
    )
    assert result is not None
    assert "螺纹" in result
    assert "35.00%" in result


def test_price_deviation_allows_when_web_reference_missing():
    result = _check_price_deviation(
        offline_price=4050,
        web_price=None,
        label="螺纹",
        config=PriceDeviationConfig(abs_tolerance=1000, pct_tolerance=0.20),
    )
    assert result is None


def test_apply_writeback_skips_price_when_deviation_is_too_large(tmp_path: Path):
    project = tmp_path / "project.xlsx"
    source = tmp_path / "source.json"
    mapping = tmp_path / "mapping.json"
    report = tmp_path / "report.json"
    _make_project(project)
    _write_json(
        source,
        {
            "meta": {"input_file": "测试钢厂报价.txt"},
            "quote_date": "2026-05-11",
            "records": [
                {"location": "蚌埠", "coil_price": 4600, "rebar_price": 4500},
            ],
        },
    )
    _write_json(
        mapping,
        [
            {
                "项目文件Sheet": "测试钢厂",
                "最新清单厂家Sheet": "测试钢厂",
                "状态": "已确认匹配",
                "说明": "",
            }
        ],
    )

    result = apply_writeback(
        project_excel=project,
        source_json_paths=[source],
        mapping_json_path=mapping,
        location="蚌埠",
        report_out=report,
    )

    wb = load_workbook(project, data_only=False)
    ws = wb["测试钢厂"]
    assert result["updated_count"] == 0
    assert result["skipped_count"] == 1
    assert "偏差过大" in result["skipped"][0]["原因"]
    assert ws["H3"].value == 3300
    assert ws["H4"].value == 3100
    wb.close()


def test_apply_writeback_allows_price_when_web_reference_is_missing(tmp_path: Path):
    project = tmp_path / "project.xlsx"
    source = tmp_path / "source.json"
    mapping = tmp_path / "mapping.json"
    report = tmp_path / "report.json"
    _make_project(project)
    wb = load_workbook(project)
    ws = wb["测试钢厂"]
    ws["G3"] = None
    ws["G4"] = None
    wb.save(project)
    wb.close()

    _write_json(
        source,
        {
            "meta": {"input_file": "测试钢厂报价.txt"},
            "quote_date": "2026-05-11",
            "records": [
                {"location": "蚌埠", "coil_price": 4600, "rebar_price": 4500},
            ],
        },
    )
    _write_json(
        mapping,
        [
            {
                "项目文件Sheet": "测试钢厂",
                "最新清单厂家Sheet": "测试钢厂",
                "状态": "已确认匹配",
                "说明": "",
            }
        ],
    )

    result = apply_writeback(
        project_excel=project,
        source_json_paths=[source],
        mapping_json_path=mapping,
        location="蚌埠",
        report_out=report,
    )

    wb = load_workbook(project, data_only=False)
    ws = wb["测试钢厂"]
    assert result["updated_count"] == 1
    assert result["skipped_count"] == 0
    assert "无网价参考" in result["updates"][0]["备注"]
    assert ws["H3"].value == 4600
    assert ws["H4"].value == 4500
    wb.close()
