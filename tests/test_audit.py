from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from ocr_price.audit import audit_image_doc_updates


def test_audit_image_doc_updates_detects_matching_values(tmp_path: Path):
    workbook = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "报价表"
    mill = wb.create_sheet("测试钢厂")
    mill["H3"] = 3400
    mill["H4"] = 3200
    wb.save(workbook)

    result = audit_image_doc_updates(
        project_excel=workbook,
        updates=[
            {
                "项目文件Sheet": "测试钢厂",
                "H3_new": 3400,
                "H4_new": 3200,
            }
        ],
    )

    assert result["status"] == "ok"
    assert result["mismatch_count"] == 0


def test_audit_image_doc_updates_reports_mismatch(tmp_path: Path):
    workbook = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "报价表"
    mill = wb.create_sheet("测试钢厂")
    mill["H3"] = 3300
    mill["H4"] = 3200
    wb.save(workbook)

    result = audit_image_doc_updates(
        project_excel=workbook,
        updates=[
            {
                "项目文件Sheet": "测试钢厂",
                "H3_new": 3400,
                "H4_new": 3200,
            }
        ],
    )

    assert result["status"] == "failed"
    assert result["mismatch_count"] == 1
    assert result["mismatches"][0]["cell"] == "H3"
