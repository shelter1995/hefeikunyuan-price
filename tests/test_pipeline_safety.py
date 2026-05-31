from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ocr_price import pipeline
from ocr_price.writeback_image_doc import apply_writeback


def _write_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _make_project(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "报价表"
    ws.cell(row=1, column=5, value="闽源")
    ws.cell(row=12, column=2, value="20")
    ws.cell(row=12, column=3, value="9")
    ws.cell(row=12, column=4, value="HRB400E")
    mill = wb.create_sheet("测试钢厂")
    mill["G3"] = 3300
    mill["G4"] = 3100
    mill["H3"] = 3300
    mill["H4"] = 3100
    wb.save(path)


def test_apply_writeback_dry_run_does_not_modify_workbook(tmp_path: Path):
    project = tmp_path / "project.xlsx"
    source = tmp_path / "source.json"
    mapping = tmp_path / "mapping.json"
    report = tmp_path / "report.json"
    _make_project(project)
    _write_json(
        source,
        {
            "meta": {"input_file": "测试钢厂报价.txt"},
            "quote_date": "2026-05-11",
            "records": [
                {"location": "蚌埠", "coil_price": 3400, "rebar_price": 3200},
            ],
            "_vision_result": {
                "库存情况": [
                    {"规格": "9米螺纹12E", "状态": "充足", "原始描述": ""},
                    {"规格": "9米螺纹12E", "状态": "告警", "原始描述": "极少"},
                ]
            },
        },
    )
    _write_json(
        mapping,
        [
            {
                "项目文件Sheet": "测试钢厂",
                "最新清单厂家Sheet": "测试钢厂",
                "状态": "已确认匹配",
                "说明": "",
            }
        ],
    )

    result = apply_writeback(
        project_excel=project,
        source_json_paths=[source],
        mapping_json_path=mapping,
        location="蚌埠",
        report_out=report,
        dry_run=True,
    )

    wb = load_workbook(project)
    ws = wb["测试钢厂"]
    assert result["dry_run"] is True
    assert result["updated_count"] == 1
    assert result["backup_file"] is None
    assert result["inventory_report"]["status"] == "review_only"
    assert result["inventory_report"]["review"]["conflict_group_count"] == 1
    assert ws["H3"].value == 3300
    assert ws["H4"].value == 3100
    wb.close()


def test_apply_writeback_inventory_uses_review_selected_item_once(tmp_path: Path):
    project = tmp_path / "project.xlsx"
    mapping = tmp_path / "mapping.json"
    report = tmp_path / "report.json"
    _make_project(project)

    text = tmp_path / "闽源.txt"
    text.write_text("9米HRB400E规格有：20（配货）", encoding="utf-8")
    fresh = tmp_path / "ocr价格提取_闽源.json"
    stale = tmp_path / "ocr价格提取_闽源_旧.json"
    _write_json(
        fresh,
        {
            "company": "闽源集团",
            "meta": {"input_file": str(text)},
            "quote_date": "2026-05-11",
            "records": [
                {"location": "蚌埠", "coil_price": 3400, "rebar_price": 3200},
            ],
        },
    )
    _write_json(
        stale,
        {
            "company": "闽源集团",
            "inventory": [
                {"规格": "9米HRB400E螺纹20", "状态": "充足", "原始描述": "旧缓存"}
            ],
        },
    )
    _write_json(
        mapping,
        [
            {
                "项目文件Sheet": "测试钢厂",
                "最新清单厂家Sheet": "闽源",
                "状态": "已确认匹配",
                "说明": "",
            }
        ],
    )

    result = apply_writeback(
        project_excel=project,
        source_json_paths=[fresh, stale],
        mapping_json_path=mapping,
        location="蚌埠",
        report_out=report,
        dry_run=False,
    )

    inventory_report = result["inventory_report"]
    assert inventory_report["status"] == "ok"
    assert inventory_report["applied_count"] == 1
    applied = inventory_report["applied"][0]
    assert applied["status"] == "告警"
    assert applied["cell"] == "E12"
    assert applied["source_spec"] == "9米HRB400E螺纹20"
    assert applied["confidence_basis"] == "原始txt解析"
    assert inventory_report["review"]["conflict_group_count"] == 1

    wb = load_workbook(project)
    ws = wb["报价表"]
    assert ws["E12"].fill.fill_type == "solid"
    assert ws["E12"].fill.start_color.rgb in {"00FFC000", "FFC000"}
    wb.close()


def test_file_sha256_changes_when_file_changes(tmp_path: Path):
    path = tmp_path / "project.xlsx"
    path.write_bytes(b"before")
    before = pipeline._file_sha256(path)

    path.write_bytes(b"after")
    after = pipeline._file_sha256(path)

    assert before != after


def test_assert_dry_run_unchanged_raises_on_modified_file(tmp_path: Path):
    path = tmp_path / "project.xlsx"
    path.write_bytes(b"before")
    before = pipeline._file_sha256(path)
    path.write_bytes(b"after")

    try:
        pipeline._assert_dry_run_unchanged(path, before)
    except RuntimeError as exc:
        assert "dry-run 修改了项目 Excel" in str(exc)
    else:
        raise AssertionError("Expected RuntimeError when dry-run changes the workbook")


def test_pipeline_lock_blocks_existing_active_lock(tmp_path: Path):
    lock_path = tmp_path / ".quote_update.lock"
    lock_path.write_text(
        json.dumps({"pid": 999999, "project": "project.xlsx"}, ensure_ascii=False),
        encoding="utf-8",
    )

    try:
        with pipeline._pipeline_lock(lock_path, project=tmp_path / "project.xlsx"):
            raise AssertionError("lock should block")
    except RuntimeError as exc:
        assert "已有报价更新任务锁" in str(exc)


def test_write_manifest_preserves_locations_and_image_sources(tmp_path: Path):
    manifest_path = tmp_path / "dry_run_manifest.json"
    artifact_dir = tmp_path / "artifacts"
    source_json = artifact_dir / "ocr价格提取_桂鑫报价.json"
    result = {
        "project": "项目报价/安徽合肥-安徽蚌埠-测试.xlsx",
        "mode": "both",
        "started_at": "2026-05-14T09:00:00",
        "ended_at": "2026-05-14T09:01:00",
        "web_location": "安徽合肥",
        "image_location": "安徽蚌埠",
        "web": {"status": "prepared", "phase": "web_prepare"},
        "image_doc": {
            "status": "prepared",
            "phase": "image_prepare",
            "source_jsons": [str(source_json)],
        },
    }

    pipeline._write_manifest(manifest_path, result, artifact_dir)

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["web_location"] == "安徽合肥"
    assert manifest["image_location"] == "安徽蚌埠"
    assert manifest["web"]["location"] == "安徽合肥"
    assert manifest["image_doc"]["location"] == "安徽蚌埠"
    assert manifest["image_doc"]["source_jsons"] == [str(source_json)]
