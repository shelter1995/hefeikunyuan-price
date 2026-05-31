"""库存颜色标注模块 - 将视觉识别的库存状态标注到Excel单元格

使用方式:
    from ocr_price.inventory_color import InventoryColorApplier

    applier = InventoryColorApplier("项目报价/XX.xlsx")
    applier.clear_all_colors()
    result = applier.apply_from_json("运行产物/ocr价格提取_徐钢报价.json")
    applier.save()
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter


# 颜色定义
COLOR_MAP = {
    "充足": "FF0070C0",  # 蓝色
    "告警": "FFFFC000",  # 黄色
    "缺货": "FFFF0000",  # 红色
}
COLOR_RESET = PatternFill(fill_type=None)


@dataclass
class InventoryItem:
    spec: str
    status: str
    raw_desc: str = ""


@dataclass
class FactoryColumnMap:
    """厂家名到Excel列的映射

    常规厂家：一个price_col
    特殊厂家（如长江）：多个extra_cols，根据仓库前缀选择
    """
    name: str
    price_col: str            # 主列（价格列）
    netdiff_col: str          # 网差列
    extra_cols: dict[str, str] = field(default_factory=dict)  # 仓库→列：{"厂内": "O", "蚌埠": "Q"}


WAREHOUSE_LABEL_TO_KEY: dict[str, str] = {
    "钢厂": "厂内",
    "厂内": "厂内",
    "蚌埠库": "蚌埠",
    "蚌埠": "蚌埠",
    "阜阳库": "阜阳",
    "阜阳": "阜阳",
    "蒙城": "蒙城",
    "合肥港": "合肥",
    "合肥铁四局": "合肥",
    "南京库": "南京",
    "安庆库": "安庆",
}


def _resolve_row9_formula_label(ws, wb, col: int) -> str:
    """Resolve row 9 formula cell to actual text label from source sheet."""
    cell = ws.cell(row=9, column=col)
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, str) and val.startswith("="):
        m = re.match(r"=(\w+)!([A-Z]+)(\d+)", val)
        if m:
            sheet_name = m.group(1)
            col_letter = m.group(2)
            src_row = int(m.group(3))
            if sheet_name in wb.sheetnames:
                src_ws = wb[sheet_name]
                col_idx = column_index_from_string(col_letter)
                src_val = src_ws.cell(row=src_row, column=col_idx).value
                return str(src_val).strip() if src_val else ""
        return ""
    return str(val).strip()


def _detect_warehouse_cols(ws, wb, mill_start_col: int) -> dict[str, str]:
    """Scan row 9 starting from mill_start_col to detect warehouse columns."""
    warehouse_cols: dict[str, str] = {}
    for c in range(mill_start_col, min(mill_start_col + 20, ws.max_column + 1)):
        label = _resolve_row9_formula_label(ws, wb, c)
        if not label:
            continue
        for label_keyword, wh_key in WAREHOUSE_LABEL_TO_KEY.items():
            if label_keyword in label:
                warehouse_cols[wh_key] = get_column_letter(c)
                break
        row1_val = ws.cell(row=1, column=c).value
        if row1_val and str(row1_val).strip() and c > mill_start_col:
            break
    return warehouse_cols


class InventoryColorApplier:
    """库存颜色标注器"""

    # 厂家名标准化映射
    FACTORY_NAME_MAP = {
        "马长江": "长江",
        "徐纲": "徐钢",
    }

    # 非厂家名排除列表
    EXCLUDE_NAMES = ["合肥鲲源", "预备发货", "钢厂", "需求单位", "提货单位"]

    def __init__(self, workbook_path: str | Path) -> None:
        self.wb = load_workbook(str(workbook_path))
        self.sheet = self.wb["报价表"]
        self._build_mappings()

    def _build_mappings(self) -> None:
        """建立厂家→列、规格→行的映射"""
        self.factory_map = self._build_factory_map()
        self.inventory_rows = self._build_inventory_row_map()

    def _build_factory_map(self) -> dict[str, FactoryColumnMap]:
        """读取第1行，建立厂家名→列映射，通过第9行动态检测仓库列"""
        factory_map: dict[str, FactoryColumnMap] = {}

        for col_idx in range(1, self.sheet.max_column + 1):
            cell_value = self.sheet.cell(row=1, column=col_idx).value
            if not cell_value:
                continue

            name = self._normalize_factory_name(cell_value)
            if not name:
                continue

            if name not in factory_map:
                col_letter = get_column_letter(col_idx)

                # 动态检测仓库列
                extra_cols = _detect_warehouse_cols(self.sheet, self.wb, col_idx)

                factory_map[name] = FactoryColumnMap(
                    name=name,
                    price_col=col_letter,
                    netdiff_col=get_column_letter(col_idx + 1),
                    extra_cols=extra_cols,
                )

        return factory_map

    def _build_inventory_row_map(self) -> dict[tuple[str, str, str | None], int]:
        """读取A-C列，建立(类型, 规格, 长度)→行号映射"""
        row_map: dict[tuple[str, str, str | None], int] = {}
        current_type: str | None = None

        for row_idx in range(9, self.sheet.max_row + 1):
            a_val = self.sheet.cell(row=row_idx, column=1).value
            b_val = self.sheet.cell(row=row_idx, column=2).value
            c_val = self.sheet.cell(row=row_idx, column=3).value

            if not a_val and not b_val:
                continue

            # 更新当前类型
            if a_val:
                a_str = str(a_val).strip()
                if "一级钢" in a_str:
                    current_type = "一级钢"
                elif "抗震" in a_str or "三级钢" in a_str:
                    current_type = "抗震三级钢"

            if not current_type or not b_val:
                continue

            spec = str(b_val).strip()
            length = str(c_val).strip() if c_val else None

            key = (current_type, spec, length)
            row_map[key] = row_idx

        return row_map

    def _normalize_factory_name(self, raw_name: str, mapping: dict[str, str] | None = None) -> str | None:
        """厂家名标准化
        
        Args:
            raw_name: 原始厂家名
            mapping: 可选的厂家对照表映射 {项目Sheet名: 来源厂家名}
        """
        name = str(raw_name).strip()

        for ex in self.EXCLUDE_NAMES:
            if ex in name:
                return None

        # 优先使用传入的厂家对照表映射
        if mapping:
            for sheet_name, source_name in mapping.items():
                if sheet_name in name or source_name in name:
                    # 返回标准化的厂家名（取来源厂家名的前2个字）
                    return source_name[:2] if len(source_name) >= 2 else source_name

        for key, val in self.FACTORY_NAME_MAP.items():
            if key in name:
                return val

        # 对于"XX-YY"格式，提取后半部分
        if "-" in name:
            parts = name.split("-")
            if len(parts) >= 2:
                return parts[-1].strip()[:2] if len(parts[-1].strip()) >= 2 else parts[-1].strip()

        # 提取核心厂家名（2-3个字）
        if len(name) >= 2:
            return name[:2]

        return name

    def _resolve_price_col(self, factory: FactoryColumnMap, inventory_spec: str) -> str:
        """根据库存规格前缀，解析应该使用哪一列"""
        if not factory.extra_cols:
            return factory.price_col

        # 检查库存规格前缀
        for prefix, col in factory.extra_cols.items():
            if prefix in inventory_spec:
                return col

        # 默认使用主列
        return factory.price_col

    def match_inventory_to_row(self, inventory_spec: str) -> int | None:
        """将库存规格映射到报价表行号"""
        numbers = re.findall(r"\d+", inventory_spec)
        if not numbers:
            return None
        spec_number = numbers[0]

        length = None
        if "9米" in inventory_spec or "9m" in inventory_spec:
            length = "9"
        elif "12米" in inventory_spec or "12m" in inventory_spec:
            length = "12"

        if "盘螺" in inventory_spec:
            steel_type = "一级钢"
        else:
            steel_type = "抗震三级钢"

        # 精确匹配
        key = (steel_type, spec_number, length)
        if key in self.inventory_rows:
            return self.inventory_rows[key]

        # 模糊匹配（无长度时）
        if length is None:
            for (stype, bspec, clen), row in self.inventory_rows.items():
                if stype == steel_type and bspec == spec_number:
                    return row

        return None

    def get_factory_mapping_report(self) -> list[dict[str, str]]:
        """生成厂家名称对照表，用于用户确认"""
        report = []
        for name, factory in self.factory_map.items():
            cols = [f"{factory.price_col}列（价格）"]
            if factory.extra_cols:
                for prefix, col in factory.extra_cols.items():
                    cols.append(f"{col}列（{prefix}库存）")
            report.append({
                "厂家名": name,
                "映射列": ", ".join(cols),
            })
        return report

    def clear_all_colors(self) -> None:
        """清空报价表所有单元格颜色"""
        for row in self.sheet.iter_rows(
            min_row=1, max_row=self.sheet.max_row, max_col=self.sheet.max_column
        ):
            for cell in row:
                if cell.fill and cell.fill.fill_type:
                    cell.fill = COLOR_RESET

    def apply_inventory_colors(
        self,
        factory_name: str,
        inventory_items: list[InventoryItem],
    ) -> dict[str, Any]:
        """对指定厂家的库存项标注颜色"""
        result = {"applied": 0, "skipped": 0, "details": []}

        factory = self.factory_map.get(factory_name)
        if not factory:
            result["skipped"] = len(inventory_items)
            return result

        for item in inventory_items:
            row = self.match_inventory_to_row(item.spec)
            if row is None:
                result["skipped"] += 1
                result["details"].append({
                    "spec": item.spec,
                    "status": item.status,
                    "reason": "报价表无对应规格",
                })
                continue

            # 根据库存前缀选择列
            price_col = self._resolve_price_col(factory, item.spec)
            cell = self.sheet[f"{price_col}{row}"]
            color = COLOR_MAP.get(item.status)
            if color:
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                result["applied"] += 1
                result["details"].append({
                    "spec": item.spec,
                    "status": item.status,
                    "cell": f"{price_col}{row}",
                })

        return result

    def apply_from_json(self, json_path: str | Path, mapping: dict[str, str] | None = None) -> dict[str, Any]:
        """从JSON文件读取库存并标注
        
        Args:
            json_path: JSON文件路径
            mapping: 可选的厂家对照表映射 {项目Sheet名: 来源厂家名}
        """
        data = json.loads(Path(json_path).read_text(encoding="utf-8"))

        company = data.get("company", "")
        vision = data.get("_vision_result", {})
        inventory_data = vision.get("库存情况", [])

        factory_name = self._normalize_factory_name(company, mapping) or company

        items = [
            InventoryItem(
                spec=item["规格"],
                status=item["状态"],
                raw_desc=item.get("原始描述", ""),
            )
            for item in inventory_data
        ]

        return self.apply_inventory_colors(factory_name, items)

    def apply_all_from_directory(self, directory: str | Path, mapping_json_path: str | Path | None = None) -> dict[str, Any]:
        """批量应用目录下所有JSON文件
        
        Args:
            directory: JSON文件目录
            mapping_json_path: 可选的厂家对照表JSON文件路径
        """
        directory = Path(directory)
        
        # 加载厂家对照表映射
        mapping = {}
        if mapping_json_path and Path(mapping_json_path).exists():
            mapping_data = json.loads(Path(mapping_json_path).read_text(encoding="utf-8"))
            for row in mapping_data:
                sheet = str(row.get("项目文件Sheet") or "").strip()
                source = str(row.get("最新清单厂家Sheet") or "").strip()
                if sheet and source:
                    mapping[sheet] = source
        
        total_result = {"applied": 0, "skipped": 0, "files": []}

        for json_file in sorted(directory.glob("ocr价格提取_*.json")):
            result = self.apply_from_json(json_file, mapping)
            total_result["applied"] += result["applied"]
            total_result["skipped"] += result["skipped"]
            total_result["files"].append({
                "file": json_file.name,
                **result,
            })

        return total_result

    def save(self, output_path: str | Path | None = None) -> None:
        """保存工作簿"""
        if output_path:
            self.wb.save(str(output_path))
        else:
            self.wb.save()


# ============ 便捷函数 ============

def update_inventory_colors(
    project_path: str | Path,
    artifact_dir: str | Path,
) -> dict[str, Any]:
    """为项目文件更新库存颜色（Pipeline集成入口）"""
    project_path = Path(project_path)
    artifact_dir = Path(artifact_dir)

    applier = InventoryColorApplier(project_path)
    applier.clear_all_colors()
    result = applier.apply_all_from_directory(artifact_dir)

    # 保存到备份文件
    backup_path = project_path.with_suffix(".backup_inventory.xlsx")
    applier.save(backup_path)

    return {
        "backup_path": str(backup_path),
        "applied_cells": result["applied"],
        "skipped_cells": result["skipped"],
        "factory_details": result["files"],
    }
