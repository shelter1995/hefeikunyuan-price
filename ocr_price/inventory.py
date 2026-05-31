from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

from .xlsx_utils import load_workbook_safe


@dataclass
class InventoryItem:
    product: str  # 螺纹, 盘螺, 线材/高线
    spec: str  # 6, 8, 10, 12, 14, etc.
    length: str | None  # 9, 12
    material: str | None  # HRB400E, HRB500E, HPB300
    status: str  # 充足, 告警, 缺货
    note: str = ""
    warehouse: str | None = None  # 蚌埠, 厂内, 阜阳, etc.
    source_file: str = ""
    source_spec: str = ""
    source_kind: str = ""
    confidence_basis: str = ""


@dataclass
class InventorySourceEntry:
    item: InventoryItem
    source_spec: str
    source_kind: str
    confidence_basis: str
    source_priority: int


# Color fills
FILL_BLUE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)

CONFIRMED_MAPPING_STATUSES = {"已确认匹配", "已确认不更新"}


SIMPLE_SPEC_RE = re.compile(
    r"(?P<spec>\d+)[eE]?\s*(?:[（(](?P<note>[^）)]+)[）)])?"
)

STATUS_ALERT_KEYWORDS = ("配货", "配", "极少", "少", "少量", "紧张")
STATUS_SHORTAGE_KEYWORDS = ("无货", "缺货", "没货", "暂无", "等生产", "停产")


def _detect_status(spec_text: str) -> tuple[str, str]:
    text = spec_text.strip()
    # Check for shortage keywords
    for kw in STATUS_SHORTAGE_KEYWORDS:
        if kw in text:
            return "缺货", kw
    # Check for alert keywords
    for kw in STATUS_ALERT_KEYWORDS:
        if kw in text:
            return "告警", kw
    # Check for quantity note like (3件), (22件), (36件)
    if re.search(r"[（(]\d+件?[）)]", text):
        note = re.search(r"[（(]([^）)]+)[）)]", text).group(1)
        return "告警", note
    # Check for numeric quantity without parentheses
    if re.search(r"\d+件", text):
        return "告警", re.search(r"(\d+件)", text).group(1)
    return "充足", ""


def _extract_specs(specs_text: str) -> list[tuple[str, str, str]]:
    """Extract (spec, status, note) tuples from specs text."""
    results: list[tuple[str, str, str]] = []
    # Split by common separators
    parts = re.split(r"[、，,；;]", specs_text)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        m = SIMPLE_SPEC_RE.match(part)
        if m:
            spec = m.group("spec")
            status, note = _detect_status(part)
            note = m.group("note") or note
            results.append((spec, status, note))
    return results


def _parse_inventory_line(line: str) -> list[InventoryItem]:
    """Parse a single line for inventory information."""
    items: list[InventoryItem] = []

    # Pattern 1: "...规格有：specs"
    # Examples: "9米HRB400E规格有：12、16、18"
    #           "铁标12米HRB400E规格有：12、14（少）、16"
    #           "12米HRB500E规格有：12（3件）、20、22"
    if "规格有" in line or "规格有" in line:
        prefix, _, specs_text = line.partition("规格有")
        specs_text = specs_text.lstrip("：:").strip()
        if not specs_text:
            return items

        # Extract length from prefix
        length = None
        m_len = re.search(r"(\d+)米", prefix)
        if m_len:
            length = m_len.group(1)

        # Extract material from prefix
        material = None
        m_mat = re.search(r"(HRB\d+E?|HPB\d+)", prefix)
        if m_mat:
            material = m_mat.group(1)

        # Extract product from prefix
        product = "螺纹"
        for p in ("盘螺", "线材", "高线"):
            if p in prefix:
                product = p
                break

        for spec, status, note in _extract_specs(specs_text):
            items.append(
                InventoryItem(
                    product=product,
                    spec=spec,
                    length=length,
                    material=material,
                    status=status,
                    note=note,
                )
            )
        return items

    # Pattern 2: "MATERIAL产品大包：specs"
    # Examples: "HRB400E盘螺大包：12（22件）、6、8"
    #           "HPB300线材大包：10"
    m = re.search(r"(HRB\d+E?|HPB\d+)?\s*(盘螺|线材|高线)(?:大包)?[：:]\s*(.+)", line)
    if m:
        material = m.group(1)
        product = m.group(2)
        specs_text = m.group(3)
        for spec, status, note in _extract_specs(specs_text):
            items.append(
                InventoryItem(
                    product=product,
                    spec=spec,
                    length=None,
                    material=material,
                    status=status,
                    note=note,
                )
            )
        return items

    # Pattern 3: "spec规格明天生产..." or "spec规格等生产"
    m = re.search(r"(\d+)(?:规格)?.*?等生产|(\d+)(?:规格)?.*?明天生产", line)
    if m:
        spec = m.group(1) or m.group(2)
        items.append(
            InventoryItem(
                product="螺纹",
                spec=spec,
                length=None,
                material=None,
                status="缺货",
                note="等生产",
            )
        )
        return items

    return items


def parse_inventory_text(text: str) -> list[InventoryItem]:
    """Parse inventory description from offline quote text."""
    items: list[InventoryItem] = []
    lines = text.splitlines()

    for line in lines:
        line = line.strip()
        if not line:
            continue
        items.extend(_parse_inventory_line(line))

    return items


def _company_match(file_company: str, target_company: str) -> bool:
    """Check if file company matches target company (fuzzy match)."""
    if not file_company or not target_company:
        return False
    # Exact match
    if file_company == target_company:
        return True
    # One contains the other
    if target_company in file_company or file_company in target_company:
        return True
    # Normalize and compare first 2 chars (common practice for Chinese company names)
    norm_file = re.sub(r"[报价钢厂集团]+$", "", file_company)
    norm_target = re.sub(r"[报价钢厂集团]+$", "", target_company)
    if norm_file == norm_target:
        return True
    if norm_target in norm_file or norm_file in norm_target:
        return True
    return False


def load_inventory_from_sources(
    source_json_paths: list[Path], company_name: str
) -> list[InventoryItem]:
    """Load inventory from source files based on company name."""
    items: list[InventoryItem] = []
    for path in source_json_paths:
        data = _load_json_safe(path)
        if not data:
            continue
        meta = data.get("meta") or {}
        input_file = str(meta.get("input_file") or "").strip()
        file_company = _extract_company_from_filename(input_file or path.name)
        if not _company_match(file_company, company_name):
            continue
        structured_items = _inventory_items_from_vision_result(data)
        if structured_items:
            items.extend(structured_items)
            continue
        # Try to read raw text from original file (pass OCR JSON for image files)
        text = _read_source_text(input_file or str(path), ocr_json=data)
        if text:
            items.extend(parse_inventory_text(text))
    return _dedupe_inventory_items(items)


def build_inventory_review(source_json_paths: list[Path]) -> dict[str, Any]:
    """Build a dry-run friendly inventory normalization and conflict report."""
    raw_items: list[dict[str, Any]] = []
    for path in source_json_paths:
        data = _load_json_safe(path)
        if not data:
            continue
        company = _source_company_from_payload(path, data)
        for entry in _raw_inventory_items_from_payload(path, data):
            raw_items.append(
                _inventory_item_review_dict(
                    entry.item,
                    company,
                    path.name,
                    entry.source_spec,
                    entry.source_kind,
                    entry.confidence_basis,
                    entry.source_priority,
                )
            )

    groups: dict[tuple[str, str, str, str, str, str], list[dict[str, Any]]] = {}
    for item in raw_items:
        key = (
            item["company"],
            item["warehouse"],
            item["product"],
            item["spec"],
            item["length"],
            item["material"],
        )
        groups.setdefault(key, []).append(item)

    selected: list[dict[str, Any]] = []
    duplicate_groups: list[dict[str, Any]] = []
    conflict_groups: list[dict[str, Any]] = []
    for group_items in groups.values():
        chosen = max(
            enumerate(group_items),
            key=lambda pair: (int(pair[1].get("source_priority") or 0), -pair[0]),
        )[1]
        selected.append(chosen)
        if len(group_items) <= 1:
            continue
        statuses = _unique_preserve_order(str(x.get("status") or "") for x in group_items)
        group = {
            "company": chosen["company"],
            "warehouse": chosen["warehouse"],
            "product": chosen["product"],
            "spec": chosen["spec"],
            "length": chosen["length"],
            "material": chosen["material"],
            "count": len(group_items),
            "statuses": statuses,
            "selected": chosen,
            "items": group_items,
        }
        duplicate_groups.append(group)
        if len(statuses) > 1:
            conflict_groups.append(group)

    return {
        "status": "ok",
        "raw_count": len(raw_items),
        "selected_count": len(selected),
        "duplicate_group_count": len(duplicate_groups),
        "conflict_group_count": len(conflict_groups),
        "selected": selected,
        "duplicate_groups": duplicate_groups,
        "conflict_groups": conflict_groups,
    }


def inventory_items_from_review(review: dict[str, Any]) -> dict[str, list[InventoryItem]]:
    """Convert selected review rows into the canonical writeback inventory map."""
    mill_inventories: dict[str, list[InventoryItem]] = {}
    for row in review.get("selected", []) or []:
        if not isinstance(row, dict):
            continue
        company = str(row.get("company") or "").strip()
        if not company:
            continue
        item = InventoryItem(
            product=str(row.get("product") or "").strip(),
            spec=str(row.get("spec") or "").strip(),
            length=str(row.get("length") or "").strip() or None,
            material=str(row.get("material") or "").strip() or None,
            status=str(row.get("status") or "").strip(),
            note=str(row.get("note") or "").strip(),
            warehouse=str(row.get("warehouse") or "").strip() or None,
            source_file=str(row.get("source_file") or "").strip(),
            source_spec=str(row.get("source_spec") or "").strip(),
            source_kind=str(row.get("source_kind") or "").strip(),
            confidence_basis=str(row.get("confidence_basis") or "").strip(),
        )
        if item.product and item.spec and item.status:
            mill_inventories.setdefault(company, []).append(item)
    return mill_inventories


def _unique_preserve_order(values: Any) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return out


def _source_company_from_payload(path: Path, data: dict[str, Any]) -> str:
    company = str(data.get("company") or "").strip()
    if company:
        return company
    vision = data.get("_vision_result")
    if isinstance(vision, dict):
        company = str(vision.get("厂家名称") or "").strip()
        if company:
            return company
    meta = data.get("meta") if isinstance(data.get("meta"), dict) else {}
    input_file = str(meta.get("input_file") or "").strip()
    return _extract_company_from_filename(input_file or path.name)


def _raw_inventory_items_from_payload(
    path: Path,
    data: dict[str, Any],
) -> list[InventorySourceEntry]:
    out: list[InventorySourceEntry] = []
    meta = data.get("meta") if isinstance(data.get("meta"), dict) else {}
    input_file = str(meta.get("input_file") or "").strip()
    input_path = Path(input_file) if input_file else None
    if input_path and input_path.suffix.lower() == ".txt" and input_path.exists():
        text = _read_source_text(input_file, ocr_json=data)
        if text:
            return [
                InventorySourceEntry(
                    item=item,
                    source_spec=_inventory_spec_label(item),
                    source_kind="original_text",
                    confidence_basis="原始txt解析",
                    source_priority=100,
                )
                for item in parse_inventory_text(text)
            ]

    containers: list[tuple[str, str, int, Any]] = []
    vision = data.get("_vision_result")
    if isinstance(vision, dict):
        containers.append(("vision_result", "MiniMax视觉库存表", 80, vision.get("库存情况")))
    containers.append(("json_inventory", "JSON库存字段", 70, data.get("库存情况")))
    containers.append(("top_level_inventory", "JSON顶层inventory", 60, data.get("inventory")))

    for source_kind, confidence_basis, source_priority, inventory in containers:
        if not isinstance(inventory, list):
            continue
        for raw_item in inventory:
            if not isinstance(raw_item, dict):
                continue
            parsed = _parse_structured_inventory_item(raw_item)
            if parsed:
                out.append(
                    InventorySourceEntry(
                        item=parsed,
                        source_spec=str(raw_item.get("规格") or "").strip(),
                        source_kind=source_kind,
                        confidence_basis=confidence_basis,
                        source_priority=source_priority,
                    )
                )
    if out:
        return out

    text = _read_source_text(input_file or str(path), ocr_json=data)
    if text:
        for item in parse_inventory_text(text):
            out.append(
                InventorySourceEntry(
                    item=item,
                    source_spec=_inventory_spec_label(item),
                    source_kind="fallback_text",
                    confidence_basis="OCR库存文本兜底解析",
                    source_priority=50,
                )
            )
    return out


def _inventory_spec_label(item: InventoryItem) -> str:
    parts = []
    if item.warehouse:
        parts.append(item.warehouse)
    if item.length:
        parts.append(f"{item.length}米")
    if item.material:
        parts.append(item.material)
    parts.append(item.product)
    parts.append(item.spec)
    return "".join(parts)


def _inventory_item_review_dict(
    item: InventoryItem,
    company: str,
    source_file: str,
    source_spec: str,
    source_kind: str,
    confidence_basis: str,
    source_priority: int,
) -> dict[str, Any]:
    return {
        "company": company,
        "source_file": source_file,
        "source_spec": source_spec or _inventory_spec_label(item),
        "source_kind": source_kind,
        "confidence_basis": confidence_basis,
        "source_priority": source_priority,
        "warehouse": _normalize_warehouse_key(item.warehouse) or "",
        "product": item.product,
        "spec": str(item.spec).strip(),
        "length": item.length or "",
        "material": item.material or "",
        "status": item.status,
        "note": item.note,
    }


def _load_json_safe(path: Path) -> dict[str, Any] | None:
    try:
        import json
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _extract_company_from_filename(filename: str) -> str:
    """Extract company name from filename or JSON data."""
    stem = Path(filename).stem
    # Remove common prefixes
    stem = re.sub(r"^ocr价格提取[_\-]?", "", stem)
    stem = re.sub(r"^test[_\-]?", "", stem)
    stem = re.sub(r"[_\-]?minimax$", "", stem)
    stem = re.sub(r"[_\-]?text$", "", stem)
    # Remove date suffixes
    stem = re.sub(r"[\-_ ]?\d{4}[\-_.]\d{1,2}[\-_.]\d{1,2}$", "", stem)
    stem = re.sub(r"[\-_ ]?\d{1,2}[\-_.]\d{1,2}$", "", stem)
    stem = re.sub(r"[\-_ ]+$", "", stem)
    return stem.strip()


def _read_source_text(input_file: str, ocr_json: dict[str, Any] | None = None) -> str | None:
    path = Path(input_file)
    if not path.exists():
        return None
    suffix = path.suffix.lower()
    if suffix == ".txt":
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix in {".jpg", ".jpeg", ".png", ".pdf", ".webp"}:
        # For image files, try to extract inventory from MiniMax vision result
        if ocr_json:
            # Try _vision_result first (MiniMax format)
            vision_result = ocr_json.get("_vision_result", {})
            if vision_result and "库存情况" in vision_result:
                inventory = vision_result["库存情况"]
                if isinstance(inventory, list):
                    lines = []
                    for raw_item in inventory:
                        if isinstance(raw_item, dict):
                            parsed = _parse_structured_inventory_item(raw_item)
                            if parsed:
                                prefix = ""
                                if parsed.warehouse:
                                    prefix += parsed.warehouse
                                if parsed.length:
                                    prefix += f"{parsed.length}米"
                                if parsed.material:
                                    prefix += parsed.material
                                line = f"{prefix}{parsed.product}规格有：{parsed.spec}"
                                if parsed.status != "充足":
                                    line += f"（{parsed.status}）"
                                lines.append(line)
                            else:
                                spec = raw_item.get("规格", "")
                                if spec:
                                    lines.append(spec)
                        elif isinstance(raw_item, str):
                            lines.append(raw_item)
                    if lines:
                        return "\n".join(lines)
                elif isinstance(inventory, str):
                    return inventory
            # Fallback: try to reconstruct from records
            records = ocr_json.get("records", [])
            if records:
                lines = []
                for record in records:
                    if isinstance(record, dict):
                        line = record.get("规格", "")
                        if record.get("状态"):
                            line += f"（{record['状态']}）"
                        if line:
                            lines.append(line)
                if lines:
                    return "\n".join(lines)
        return None
    return None


def _inventory_items_from_vision_result(data: dict[str, Any]) -> list[InventoryItem]:
    vision_result = data.get("_vision_result")
    if not isinstance(vision_result, dict):
        return []
    inventory = vision_result.get("库存情况")
    if not isinstance(inventory, list):
        return []

    items: list[InventoryItem] = []
    for raw_item in inventory:
        if not isinstance(raw_item, dict):
            continue
        parsed = _parse_structured_inventory_item(raw_item)
        if parsed:
            items.append(parsed)
    return items


WAREHOUSE_KEYWORDS_INV = (
    "蚌埠库", "蚌埠", "钢厂", "厂内", "场内", "阜阳库", "阜阳", "蒙城", "合肥港", "合肥铁四局", "南京库", "安庆库",
)


def _extract_warehouse(spec_text: str) -> str | None:
    """Extract warehouse location from spec text."""
    m = re.search(rf"\(([^)]*(?:{'|'.join(WAREHOUSE_KEYWORDS_INV)})[^)]*)\)", spec_text)
    if m:
        return m.group(1).strip()
    m = re.match(rf"({'|'.join(WAREHOUSE_KEYWORDS_INV)})", spec_text)
    if m:
        return m.group(1)
    return None


def _extract_material(spec_text: str) -> str | None:
    """Extract material grade from spec text, e.g. HRB400E, HRB500E, HPB300."""
    m = re.search(r"(HRB\d+E?|HPB\d+)", spec_text, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    return None


def _extract_spec_number(spec_text: str, product: str) -> str | None:
    tail = spec_text
    if product in tail:
        tail = tail.split(product, 1)[1]
    tail = re.sub(r"\([^)]*\)|（[^）]*）", " ", tail)
    tail = re.sub(r"(HRB|HPB)\d+E?", " ", tail, flags=re.IGNORECASE)
    tail = re.sub(r"\d+\s*[米mM]", " ", tail)
    m = re.search(r"(\d+)[eE]?", tail)
    if m:
        return m.group(1)

    cleaned = re.sub(r"(HRB|HPB)\d+E?", " ", spec_text, flags=re.IGNORECASE)
    cleaned = re.sub(r"\d+\s*[米mM]", " ", cleaned)
    numbers = re.findall(r"\d+", cleaned)
    if numbers:
        return numbers[-1]
    return None


def _parse_structured_inventory_item(item: dict[str, Any]) -> InventoryItem | None:
    spec_text = str(item.get("规格") or "").strip()
    if not spec_text:
        return None

    status = str(item.get("状态") or "").strip()
    note = str(item.get("原始描述") or "").strip()
    if status not in {"充足", "告警", "缺货"}:
        status, detected_note = _detect_status(note or spec_text)
        note = note or detected_note

    length = None
    m_len = re.search(r"(\d+)\s*[米mM]", spec_text)
    if m_len:
        length = m_len.group(1)

    product = "螺纹"
    for candidate in ("盘螺", "线材", "高线", "圆钢", "螺纹"):
        if candidate in spec_text:
            product = candidate
            break
    if product == "高线":
        product = "线材"

    spec = _extract_spec_number(spec_text, product)
    if not spec:
        return None

    material = _extract_material(spec_text)
    warehouse = _extract_warehouse(spec_text)

    return InventoryItem(
        product=product,
        spec=spec,
        length=length,
        material=material,
        status=status,
        note=note,
        warehouse=warehouse,
    )


def _norm_text(s: str) -> str:
    return re.sub(r"\s+", "", s).upper()


def _match_product(item_product: str, row_product: str) -> bool:
    """Match inventory product with row product type."""
    item_norm = _norm_text(item_product)
    row_norm = _norm_text(row_product)
    if "螺纹" in item_norm and "螺纹" in row_norm:
        return True
    if "盘螺" in item_norm and "盘螺" in row_norm:
        return True
    if ("线材" in item_norm or "高线" in item_norm) and ("线材" in row_norm or "高线" in row_norm):
        return True
    return item_norm == row_norm


def _match_material(item_material: str | None, row_material: str) -> bool:
    if not item_material:
        # 无材质信息时通配匹配（兼容旧数据），但不匹配明显不同材质的行
        # 如果行材质明确标注了HRB500E等，跳过
        row_norm = _norm_text(row_material or "")
        if "HRB500" in row_norm or "HRB600" in row_norm:
            return False
        return True
    item_norm = _norm_text(item_material)
    row_norm = _norm_text(row_material or "")
    if not row_norm:
        return True
    return item_norm == row_norm or item_norm in row_norm or row_norm in item_norm


def _match_length(item_length: str | None, row_length: str | None) -> bool:
    if not item_length:
        return True
    return item_length == (row_length or "")


def _match_spec(item_spec: str, row_spec: str) -> bool:
    return item_spec == str(row_spec).strip()


def _detect_product_type(row: int, ws: Any) -> str:
    """Detect product type from A column (等级)."""
    length = ws.cell(row=row, column=3).value
    spec = ws.cell(row=row, column=2).value
    val = ws.cell(row=row, column=1).value
    if val:
        text = str(val).upper()
        if "一级钢" in text:
            return "线材"  # 一级钢 usually corresponds to 高线/线材
    # Check D column material
    mat = ws.cell(row=row, column=4).value
    if mat:
        mat_text = str(mat).upper()
        if "HPB" in mat_text:
            return "线材"
        if "HRB" in mat_text and not length and str(spec or "").strip() in {"6", "8", "10", "12"}:
            return "盘螺"
    # Default based on row context: rows 9-11 are 一级钢 (line), 12+ are 抗震三级钢 (rebar)
    if row <= 11:
        return "线材"
    return "螺纹"


WAREHOUSE_LABEL_TO_KEY: dict[str, str] = {
    "场内": "厂内",
    "钢厂": "厂内",
    "厂内": "厂内",
    "蚌埠库": "蚌埠",
    "蚌埠": "蚌埠",
    "阜阳库": "阜阳",
    "阜阳": "阜阳",
    "蒙城": "蒙城",
    "合肥港": "合肥",
    "合肥铁四局": "合肥",
    "南京库": "南京",
    "安庆库": "安庆",
}

WAREHOUSE_ALLOWED_BY_MILL: dict[str, set[str]] = {
    "长江": {"厂内", "蚌埠"},
}


def _normalize_warehouse_key(raw: str | None) -> str | None:
    if not raw:
        return None
    text = str(raw).strip()
    for label_keyword, wh_key in WAREHOUSE_LABEL_TO_KEY.items():
        if label_keyword in text:
            return wh_key
    return text


def _allowed_warehouses_for_mill(mill_name: str) -> set[str] | None:
    norm = _norm_text(mill_name)
    for key, allowed in WAREHOUSE_ALLOWED_BY_MILL.items():
        if key in norm:
            return allowed
    return None


def _dedupe_inventory_items(items: list[InventoryItem]) -> list[InventoryItem]:
    deduped: list[InventoryItem] = []
    seen: set[tuple[str, str, str | None, str | None, str | None]] = set()
    for item in items:
        key = (
            _norm_text(item.product),
            str(item.spec).strip(),
            item.length,
            _norm_text(item.material or "") or None,
            _normalize_warehouse_key(item.warehouse),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


def _resolve_row9_formula_label(ws: Any, wb: Any, col: int) -> str:
    """Resolve row 9 formula cell to actual text label from source sheet."""
    cell = ws.cell(row=9, column=col)
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, str) and val.startswith("="):
        m = re.match(r"=(\w+)!([A-Z]+)(\d+)", val)
        if m:
            sheet_name = m.group(1)
            col_letter = m.group(2)
            src_row = int(m.group(3))
            if sheet_name in wb.sheetnames:
                src_ws = wb[sheet_name]
                col_idx = column_index_from_string(col_letter)
                src_val = src_ws.cell(row=src_row, column=col_idx).value
                return str(src_val).strip() if src_val else ""
        return ""
    return str(val).strip()


def _build_warehouse_col_map(ws: Any, wb: Any, mill_start_col: int) -> dict[str, int]:
    """Build warehouse→column mapping for a multi-warehouse mill by scanning row 9."""
    warehouse_cols: dict[str, int] = {}
    for c in range(mill_start_col, min(mill_start_col + 20, ws.max_column + 1)):
        label = _resolve_row9_formula_label(ws, wb, c)
        if not label:
            continue
        for label_keyword, wh_key in WAREHOUSE_LABEL_TO_KEY.items():
            if label_keyword in label:
                warehouse_cols[wh_key] = c
                break
        row1_val = ws.cell(row=1, column=c).value
        if row1_val and str(row1_val).strip() and c > mill_start_col:
            break
    return warehouse_cols


def apply_inventory_to_project(
    project_excel: Path,
    mill_inventories: dict[str, list[InventoryItem]],
    sheet_name: str = "报价表",
    mapping_json_path: Path | None = None,
    clear_existing_colors: bool = False,
) -> dict[str, Any]:
    """
    Apply inventory colors to the quote sheet.
    mill_inventories: {mill_name: [InventoryItem, ...]}
    """
    wb = load_workbook_safe(project_excel)
    if sheet_name not in wb.sheetnames:
        return {"status": "skipped", "reason": f"{sheet_name} not found"}

    ws = wb[sheet_name]

    # Build mill -> column mapping from row 1 and row 8
    mill_to_col: dict[str, int] = {}
    for row in (1, 8):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str):
                mill = val.strip()
                if mill and mill not in (
                    "钢厂",
                    "预备发货厂家",
                    "合肥鲲源贸易有限公司钢材报价单",
                    "等级",
                    "规格(mm)",
                    "长度（米）",
                    "材质",
                    "数量",
                    "金额",
                    "提货网差",
                    "数量（吨）",
                    "网差",
                ):
                    mill_to_col[mill] = col

    # Build warehouse→column mapping for multi-warehouse mills by scanning row 9
    mill_warehouse_cols: dict[str, dict[str, int]] = {}
    for mill_name, start_col in list(mill_to_col.items()):
        wcols = _build_warehouse_col_map(ws, wb, start_col)
        if wcols:
            norm = _norm_text(mill_name)
            mill_warehouse_cols[norm] = wcols

    # Build row mapping
    row_map: list[dict[str, Any]] = []
    for r in range(9, ws.max_row + 1):
        spec = ws.cell(row=r, column=2).value
        length = ws.cell(row=r, column=3).value
        material = ws.cell(row=r, column=4).value
        if spec is None:
            continue
        product = _detect_product_type(r, ws)
        row_map.append({
            "row": r,
            "spec": str(spec).strip(),
            "length": str(length).strip() if length else None,
            "material": str(material).strip() if material else None,
            "product": product,
        })

    confirmed_mapping = _load_confirmed_mill_mapping(mapping_json_path)
    source_to_sheet_mill = {
        source_mill: _resolve_sheet_mill(source_mill, confirmed_mapping)
        for source_mill in mill_inventories
    }

    cleared_count = 0
    if clear_existing_colors and row_map:
        target_cols: set[int] = set()
        for resolved_mill in source_to_sheet_mill.values():
            norm_resolved = _norm_text(resolved_mill)
            for mapped_mill, col in mill_to_col.items():
                if _mill_match(mapped_mill, resolved_mill):
                    target_cols.add(col)
                    wh_cols = mill_warehouse_cols.get(_norm_text(mapped_mill), {})
                    if not wh_cols:
                        wh_cols = mill_warehouse_cols.get(norm_resolved, {})
                    for wh_col in wh_cols.values():
                        target_cols.add(wh_col)
                    break

        for row_info in row_map:
            row = row_info["row"]
            for col in target_cols:
                ws.cell(row=row, column=col).fill = CLEAR_FILL
                cleared_count += 1

    applied: list[dict[str, Any]] = []
    applied_cells: set[tuple[str, int, int]] = set()

    for source_mill, items in mill_inventories.items():
        resolved_mill = source_to_sheet_mill.get(source_mill, source_mill)
        norm_resolved = _norm_text(resolved_mill)
        # Find column for this mill
        target_col = None
        matched_mill = None
        for mapped_mill, col in mill_to_col.items():
            if _mill_match(mapped_mill, resolved_mill):
                target_col = col
                matched_mill = mapped_mill
                break

        if target_col is None:
            continue

        wh_cols = mill_warehouse_cols.get(_norm_text(matched_mill or ""), {})
        if not wh_cols:
            wh_cols = mill_warehouse_cols.get(norm_resolved, {})
        allowed_warehouses = _allowed_warehouses_for_mill(resolved_mill)
        if allowed_warehouses is None and matched_mill:
            allowed_warehouses = _allowed_warehouses_for_mill(matched_mill)

        for item in items:
            # Determine target column: use warehouse-specific column if available
            item_col = target_col
            if item.warehouse and wh_cols:
                warehouse_key = _normalize_warehouse_key(item.warehouse)
                if allowed_warehouses is not None and warehouse_key not in allowed_warehouses:
                    continue
                mapped_col = wh_cols.get(warehouse_key or item.warehouse)
                if mapped_col:
                    item_col = mapped_col
                elif item.warehouse in WAREHOUSE_LABEL_TO_KEY:
                    key = WAREHOUSE_LABEL_TO_KEY[item.warehouse]
                    mapped_col = wh_cols.get(key)
                    if mapped_col:
                        item_col = mapped_col
                    else:
                        # Warehouse not in this mill's columns, skip
                        continue
                else:
                    continue

            # Find matching row
            for row_info in row_map:
                if (
                    _match_spec(item.spec, row_info["spec"])
                    and _match_material(item.material, row_info["material"] or "")
                    and _match_length(item.length, row_info["length"])
                    and _match_product(item.product, row_info["product"])
                ):
                    applied_key = (_norm_text(resolved_mill), row_info["row"], item_col)
                    if applied_key in applied_cells:
                        break
                    applied_cells.add(applied_key)
                    cell = ws.cell(row=row_info["row"], column=item_col)
                    if item.status == "充足":
                        cell.fill = FILL_BLUE
                    elif item.status == "告警":
                        cell.fill = FILL_YELLOW
                    elif item.status == "缺货":
                        cell.fill = FILL_RED
                    applied.append({
                        "mill": source_mill,
                        "sheet_mill": resolved_mill,
                        "row": row_info["row"],
                        "col": item_col,
                        "cell": ws.cell(row=row_info["row"], column=item_col).coordinate,
                        "product": item.product,
                        "spec": item.spec,
                        "length": item.length,
                        "material": item.material,
                        "warehouse": item.warehouse,
                        "status": item.status,
                        "source_file": item.source_file,
                        "source_spec": item.source_spec,
                        "source_kind": item.source_kind,
                        "confidence_basis": item.confidence_basis,
                    })
                    break

    wb.save(project_excel)
    return {
        "status": "ok",
        "applied_count": len(applied),
        "cleared_count": cleared_count,
        "applied": applied,
    }


def _mill_match(mapped: str, target: str) -> bool:
    """Check if mapped mill name matches target mill name."""
    mapped_norm = _norm_text(mapped)
    target_norm = _norm_text(target)
    if mapped_norm == target_norm:
        return True
    if mapped_norm in target_norm or target_norm in mapped_norm:
        return True
    return False


def _load_confirmed_mill_mapping(mapping_json_path: Path | None) -> dict[str, str]:
    """Load confirmed source->sheet mill mapping from mapping json."""
    if not mapping_json_path:
        return {}

    data = _load_json_safe(mapping_json_path)
    if not isinstance(data, list):
        return {}

    mapping: dict[str, str] = {}
    for row in data:
        if not isinstance(row, dict):
            continue
        status = str(row.get("状态") or "").strip()
        if status not in CONFIRMED_MAPPING_STATUSES:
            continue

        source_mill = str(row.get("最新清单厂家Sheet") or "").strip()
        sheet_mill = str(row.get("项目文件Sheet") or "").strip()
        if not source_mill or not sheet_mill:
            continue

        key = _norm_text(source_mill)
        old = mapping.get(key)
        if old is None or _mill_match(old, sheet_mill):
            mapping[key] = sheet_mill

    return mapping


def _resolve_sheet_mill(source_mill: str, confirmed_mapping: dict[str, str]) -> str:
    return confirmed_mapping.get(_norm_text(source_mill), source_mill)
