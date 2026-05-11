from __future__ import annotations

import io
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


def _patch_styles_xml(style_xml: bytes) -> tuple[bytes, bool]:
    root = ET.fromstring(style_xml)
    if not root.tag.startswith("{"):
        return style_xml, False
    ns_uri = root.tag.split("}", 1)[0][1:]
    fills = root.find(f"{{{ns_uri}}}fills")
    if fills is None:
        return style_xml, False

    changed = False
    for fill in list(fills):
        if len(list(fill)) == 0:
            ET.SubElement(fill, f"{{{ns_uri}}}patternFill")
            changed = True

    if not changed:
        return style_xml, False
    return ET.tostring(root, encoding="utf-8", xml_declaration=True), True


def _build_repaired_stream(path: Path) -> io.BytesIO | None:
    changed = False
    buf = io.BytesIO()
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(buf, "w") as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "xl/styles.xml":
                data, patched = _patch_styles_xml(data)
                changed = changed or patched
            zout.writestr(info, data)
    if not changed:
        return None
    buf.seek(0)
    return buf


def load_workbook_safe(path: Path | str, **kwargs):
    try:
        return load_workbook(path, **kwargs)
    except TypeError as exc:
        message = str(exc)
        if "openpyxl.styles.fills.Fill" not in message:
            raise
        stream = _build_repaired_stream(Path(path))
        if stream is None:
            raise
        return load_workbook(stream, **kwargs)

