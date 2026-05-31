from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.styles import Font

from .inventory import (
    apply_inventory_to_project,
    build_inventory_review,
    inventory_items_from_review,
)
from .offline_validation import validate_offline_payload
from .rules import (
    CONFIRMED_SKIP_STATUS,
    CONFIRMED_WRITE_STATUS,
    PriceDeviationConfig,
    check_price_deviation as _check_price_deviation,
    coerce_price as _coerce_price,
)
from .semantic_adjustment import (
    SemanticAdjustmentError,
    interpret_supplement_adjustment_with_llm,
)
from .xlsx_utils import load_workbook_safe


def _read_source_text(input_file: str, ocr_json: dict[str, Any] | None = None) -> str | None:
    path = Path(input_file)
    if not path.exists():
        return None
    suffix = path.suffix.lower()
    if suffix == ".txt":
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix in {".jpg", ".jpeg", ".png", ".pdf", ".webp"} and ocr_json:
        vision = ocr_json.get("_vision_result")
        if isinstance(vision, dict) and isinstance(vision.get("库存情况"), list):
            lines: list[str] = []
            for item in vision.get("库存情况", []):
                if not isinstance(item, dict):
                    continue
                spec = str(item.get("规格") or "").strip()
                status = str(item.get("状态") or "").strip()
                note = str(item.get("原始描述") or "").strip()
                if not spec:
                    continue
                marker = note or status
                lines.append(f"{spec}（{marker}）" if marker else spec)
            if lines:
                return "\n".join(lines)
    return None


SKIP_SHEETS = {"报价表"}
MAPPING_HEADERS = ["项目文件Sheet", "最新清单厂家Sheet", "状态", "说明"]
MANUFACTURER_STOPWORDS = (
    "安徽",
    "江苏",
    "河南",
    "集团",
    "钢铁",
    "钢厂",
    "贸易",
    "有限公司",
    "有限责任公司",
    "公司",
    "报价",
    "补充",
    "价格",
    "提取",
    "ocr",
)


@dataclass
class SourcePrice:
    company: str
    source_file: str
    quote_date: str
    location: str
    rebar_price: int | None
    coil_price: int | None
    is_adjustment: bool = False  # True if price is an adjustment value, not absolute price
    base_source: str | None = None  # Reference to base price source if adjustment
    is_electronic_negotiation: bool = False  # True if price is "电议" (electronic negotiation)


def _extract_company_from_filename(filename: str) -> str:
    """
    Extract manufacturer from file name, e.g.:
    - 闽曜4.13.jpg -> 闽曜
    - ocr价格提取_闽源4.13.json -> 闽源
    """
    stem = Path(filename).stem
    stem = re.sub(r"^ocr价格提取[_\-]?", "", stem)
    # Remove common trailing date suffixes: 4.13 / 2026-04-13 / 2026_04_13
    stem = re.sub(r"[\-_ ]?\d{4}[\-_.]\d{1,2}[\-_.]\d{1,2}$", "", stem)
    stem = re.sub(r"[\-_ ]?\d{1,2}[\-_.]\d{1,2}$", "", stem)
    stem = re.sub(r"[\-_ ]+$", "", stem)
    return stem.strip()


def _normalize_company(name: str) -> str:
    value = re.sub(r"\s+", "", name or "")
    # 常见输入/文件名错别字：报价图片常写“徐刚”，项目 sheet 通常是“徐钢”。
    value = value.replace("刚", "钢")
    for token in MANUFACTURER_STOPWORDS:
        value = value.replace(token, "")
    return value


def _score_company_match(left: str, right: str) -> int:
    l = _normalize_company(left)
    r = _normalize_company(right)
    if not l or not r:
        return 0
    if l == r:
        return 100
    if l.startswith(r) or r.startswith(l):
        return 80
    if l in r or r in l:
        return 60
    if len(l) >= 2 and len(r) >= 2 and l[:2] == r[:2]:
        return 30
    return 0


def _resolve_quote_date(raw_date: str | None, source_file: str) -> str:
    now_year = datetime.now().year
    if raw_date:
        raw = raw_date.strip()
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", raw)
        if m:
            return raw
        m = re.match(r"^(\d{1,2})-(\d{1,2})$", raw)
        if m:
            return f"{now_year:04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    m = re.search(r"(?<!\d)(\d{1,2})[.\-_](\d{1,2})(?!\d)", Path(source_file).stem)
    if m:
        return f"{now_year:04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return datetime.now().strftime("%Y-%m-%d")


def _norm_location(text: str) -> str:
    val = re.sub(r"\s+", "", text or "")
    for token in ("省", "市", "地区", "市场", "区域", "报价"):
        val = val.replace(token, "")
    return val


def _location_match_score(target: str, rec: dict[str, Any]) -> int:
    if not target:
        return 1
    loc = _norm_location(str(rec.get("location") or ""))
    region = _norm_location(str(rec.get("region_title") or ""))
    if target and target in loc:
        return 100
    if target and target in region:
        return 90
    if loc and (loc in target or target in loc):
        return 70
    if region and (region in target or target in region):
        return 60
    return 0


def _pick_record_for_location(payload: dict[str, Any], location: str) -> dict[str, Any] | None:
    records = payload.get("records") or []
    if not isinstance(records, list):
        return None
    target = _norm_location(location.strip())

    candidates: list[tuple[int, int, int, dict[str, Any]]] = []
    for idx, rec in enumerate(records):
        if not isinstance(rec, dict):
            continue
        score = _location_match_score(target, rec)
        has_prices = int(rec.get("rebar_price") is not None and rec.get("coil_price") is not None)
        candidates.append((score, has_prices, -idx, rec))

    if not candidates:
        return None

    if target:
        matched = [x for x in candidates if x[0] > 0]
        if not matched:
            return None
        matched.sort(reverse=True)
        return matched[0][3]

    candidates.sort(reverse=True)
    return candidates[0][3]


def _extract_rebar_14_from_text(text: str) -> int | None:
    """Try to extract Φ14 rebar price from raw text."""
    # Patterns like: "螺纹14 3160", "14规格 3160", "Φ14 3160", "14# 3160"
    patterns = [
        r"(?:螺纹|螺纹钢)\s*14\D{0,10}(\d{3,5})",
        r"14\s*(?:规格|#|号)\D{0,10}(\d{3,5})",
        r"[Φφ]14\D{0,10}(\d{3,5})",
        r"14\s*[:：]\s*(\d{3,5})",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            price = int(m.group(1))
            if 1000 <= price <= 10000:
                return price
    return None


def _missing_reference_notes(ws: Any, src: SourcePrice) -> list[str]:
    notes: list[str] = []
    if src.coil_price is not None and _coerce_price(ws["G3"].value) is None:
        notes.append("盘螺无网价参考")
    if src.rebar_price is not None and _coerce_price(ws["G4"].value) is None:
        notes.append("螺纹无网价参考")
    return notes


def _load_single_source_price(path: Path, location: str) -> SourcePrice | None:
    """Load a single source price from a JSON file."""
    data = json.loads(path.read_text(encoding="utf-8"))
    validation = validate_offline_payload(data, target_location=location)
    if not validation.is_valid:
        return None

    rec = _pick_record_for_location(data, location=location)
    meta = data.get("meta") if isinstance(data.get("meta"), dict) else {}
    input_file = str(meta.get("input_file") or "").strip()
    company = _extract_company_from_filename(input_file or path.name)
    if not company:
        return None
    quote_date = _resolve_quote_date(str(data.get("quote_date") or "").strip() or None, str(path))

    rebar_price = rec.get("rebar_price") if rec else None
    coil_price = rec.get("coil_price") if rec else None
    rec_location = str(rec.get("location") or location) if rec else location
    is_electronic = False

    # If this is a supplement text with adjustment info, try to compute final price
    computed = None
    if rebar_price is None and coil_price is None:
        computed = _compute_prices_from_supplement(data, location, cache_path=path)
        if computed:
            rebar_price = computed.get("rebar_price")
            coil_price = computed.get("coil_price")
            is_electronic = bool(computed.get("is_electronic_negotiation"))
    is_adj = bool(computed and computed.get("is_adjustment"))

    # Special handling for 金虹: try to extract rebar Φ14 price if available
    norm_company = _normalize_company(company)
    if "金虹" in norm_company or "金虹" in company:
        raw_text = _read_source_text(input_file or str(path), ocr_json=data)
        if raw_text:
            price_14 = _extract_rebar_14_from_text(raw_text)
            if price_14 is not None:
                rebar_price = price_14
                is_adj = False  # Override adjustment if direct price found

    if rebar_price is None and coil_price is None and not is_adj and not is_electronic:
        return None

    return SourcePrice(
        company=company,
        source_file=path.name,
        quote_date=quote_date,
        location=rec_location,
        rebar_price=rebar_price,
        coil_price=coil_price,
        is_adjustment=is_adj,
        is_electronic_negotiation=is_electronic,
    )


def load_source_prices(source_json_paths: list[Path], location: str) -> list[SourcePrice]:
    """Load source prices, merging base prices with adjustments for the same mill.
    
    Priority rules:
    1. Supplement text (补充) takes priority over image base prices
    2. Electronic negotiation (电议) is preserved and reported
    3. Adjustments are merged with base prices when possible
    """
    # First pass: collect all source prices
    all_prices: list[SourcePrice] = []
    for path in source_json_paths:
        try:
            sp = _load_single_source_price(path, location)
            if sp:
                all_prices.append(sp)
        except Exception:
            continue

    # Group by normalized company name
    by_company: dict[str, list[SourcePrice]] = {}
    for sp in all_prices:
        key = _normalize_company(sp.company)
        by_company.setdefault(key, []).append(sp)

    # Second pass: merge and prioritize
    output: list[SourcePrice] = []
    for key, items in by_company.items():
        # Separate by type
        electronic_items = [x for x in items if x.is_electronic_negotiation]
        adj_items = [x for x in items if x.is_adjustment]
        normal_items = [x for x in items if not x.is_adjustment and not x.is_electronic_negotiation]
        
        # Check if any supplement file exists for this company
        has_supplement = any("补充" in x.source_file for x in items)
        
        if electronic_items:
            # If supplement says "电议", report it
            output.append(electronic_items[0])
        elif adj_items and normal_items:
            # Merge adjustment with base price
            base_item = normal_items[0]
            adj = adj_items[0]
            final_rebar = None
            final_coil = None
            if base_item.rebar_price is not None and adj.rebar_price is not None:
                final_rebar = base_item.rebar_price + adj.rebar_price
            if base_item.coil_price is not None and adj.coil_price is not None:
                final_coil = base_item.coil_price + adj.coil_price
            output.append(
                SourcePrice(
                    company=base_item.company,
                    source_file=f"{base_item.source_file}+{adj.source_file}",
                    quote_date=base_item.quote_date,
                    location=base_item.location,
                    rebar_price=final_rebar,
                    coil_price=final_coil,
                    is_adjustment=False,
                )
            )
        elif adj_items and has_supplement:
            # Supplement exists but no base price - use adjustment as-is
            # This means supplement has direct price or adjustment without base
            output.append(adj_items[0])
        elif normal_items:
            # Prioritize supplement files over regular files
            supplement_normals = [x for x in normal_items if "补充" in x.source_file]
            if supplement_normals:
                output.append(supplement_normals[0])
            else:
                output.append(normal_items[0])
        elif adj_items:
            output.append(adj_items[0])

    return output


def _semantic_cache_key(location: str) -> str:
    return _norm_location(location) or location


def _semantic_result_from_cache(data: dict[str, Any], location: str) -> dict[str, Any] | None:
    cache = data.get("_semantic_adjustments")
    if not isinstance(cache, dict):
        return None
    cached = cache.get(_semantic_cache_key(location)) or cache.get(location)
    return cached if isinstance(cached, dict) else None


def _write_semantic_result_cache(
    cache_path: Path | None,
    data: dict[str, Any],
    location: str,
    result: dict[str, Any],
) -> None:
    if cache_path is None:
        return
    data.setdefault("_semantic_adjustments", {})[_semantic_cache_key(location)] = result
    try:
        cache_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except OSError:
        pass


def _compute_prices_from_supplement(
    data: dict[str, Any],
    location: str,
    cache_path: Path | None = None,
) -> dict[str, Any] | None:
    """
    Try to compute prices from supplement text that contains adjustments like:
    '蚌埠下30', '阜阳下30', '合肥电议', etc.
    """
    records = data.get("records") or []
    if records:
        return None
    meta = data.get("meta") or {}
    input_file = str(meta.get("input_file") or "").strip()
    # Only process supplement texts
    if "补充" not in input_file and "supplement" not in input_file.lower():
        return None
    # Try to read the original text file
    path = Path(input_file)
    if not path.exists():
        return None
    text = path.read_text(encoding="utf-8", errors="ignore")
    cached = _semantic_result_from_cache(data, location)
    if cached:
        return cached

    company = _extract_company_from_filename(input_file)
    try:
        semantic = interpret_supplement_adjustment_with_llm(
            text=text,
            location=location,
            company=company,
        )
    except SemanticAdjustmentError:
        semantic = None
    if semantic:
        _write_semantic_result_cache(cache_path, data, location, semantic)
        return semantic

    return _parse_adjustment_text(text, location)


def _parse_adjustment_text(text: str, location: str) -> dict[str, Any] | None:
    """
    Parse text like:
    '阜阳下30，蚌埠下30，合肥电议！山东，江苏，安徽，浙江，湖北公共区域下35'
    Also handles direct prices like:
    '蚌埠：3195螺纹、3455盘螺'
    Returns {"rebar_price": int, "coil_price": int, "is_adjustment": bool} or None.
    Also handles "电议" (electronic negotiation) case.
    """
    target = _norm_location(location)
    lines = text.splitlines()
    rebar_adj = None
    coil_adj = None
    is_electronic = False
    direct_rebar = None
    direct_coil = None

    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Check for "电议" first - but only for the specific city, not partial matches
        # Use word boundary to avoid matching "蚌埠" when looking for "合肥电议"
        if re.search(rf"{re.escape(target)}(?:\s*[:：])?\s*电议", line):
            is_electronic = True
            break
        # Match patterns like '蚌埠下30' or '合肥下20' or '南京+10'
        m = re.search(rf"{re.escape(target)}\s*([上下+])\s*(\d+)", line)
        if m:
            op = m.group(1)
            val = int(m.group(2))
            rebar_adj = val if op == "+" or op == "上" else -val
            coil_adj = rebar_adj
            break
        # Match direct price patterns like '蚌埠：3195螺纹、3455盘螺'
        m_direct = re.search(
            rf"{re.escape(target)}[：:]\s*(?P<rebar>\d{{3,5}})(?:螺纹|螺纹钢)\D*(?P<coil>\d{{3,5}})(?:盘螺|盘圆)",
            line,
        )
        if m_direct:
            direct_rebar = int(m_direct.group("rebar"))
            direct_coil = int(m_direct.group("coil"))
            break

    if is_electronic:
        return {
            "rebar_price": None,
            "coil_price": None,
            "is_adjustment": False,
            "is_electronic_negotiation": True,
        }

    if direct_rebar is not None or direct_coil is not None:
        return {
            "rebar_price": direct_rebar,
            "coil_price": direct_coil,
            "is_adjustment": False,
            "is_direct_price": True,
        }

    if rebar_adj is None:
        return None

    return {
        "rebar_price": rebar_adj,
        "coil_price": coil_adj,
        "is_adjustment": True,
    }


def _write_mapping_json_csv(mapping_rows: list[dict[str, str]], json_path: Path, csv_path: Path) -> None:
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(mapping_rows, ensure_ascii=False, indent=2), encoding="utf-8")

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MAPPING_HEADERS)
        writer.writeheader()
        for row in mapping_rows:
            writer.writerow({k: row.get(k, "") for k in MAPPING_HEADERS})


def _mapping_status_priority(status: str) -> int:
    s = str(status or "").strip()
    if s == CONFIRMED_WRITE_STATUS:
        return 50
    if s == CONFIRMED_SKIP_STATUS:
        return 40
    if s.startswith("待确认"):
        return 30
    if s.startswith("未匹配"):
        return 20
    if s.startswith("跳过"):
        return 10
    return 0


def _mapping_row_key(row: dict[str, str]) -> tuple[str, str] | None:
    sheet = str(row.get("项目文件Sheet") or "").strip()
    source = str(row.get("最新清单厂家Sheet") or "").strip()
    if sheet:
        return ("sheet", sheet)
    if source:
        norm = _normalize_company(source) or source
        return ("source", norm)
    return None


def _mapping_row_score(row: dict[str, str], idx: int) -> tuple[int, int, int, int]:
    source = str(row.get("最新清单厂家Sheet") or "").strip()
    note = str(row.get("说明") or "").strip()
    return (_mapping_status_priority(str(row.get("状态") or "")), 1 if source else 0, 1 if note else 0, idx)


def _dedupe_mapping_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    if not rows:
        return rows
    best_idx_by_key: dict[tuple[str, str], int] = {}
    best_score_by_key: dict[tuple[str, str], tuple[int, int, int, int]] = {}
    keep_nonkey: set[int] = set()
    for idx, row in enumerate(rows):
        if not isinstance(row, dict):
            keep_nonkey.add(idx)
            continue
        key = _mapping_row_key(row)
        if key is None:
            keep_nonkey.add(idx)
            continue
        score = _mapping_row_score(row, idx)
        prev_score = best_score_by_key.get(key)
        if prev_score is None or score > prev_score:
            best_score_by_key[key] = score
            best_idx_by_key[key] = idx
    keep_idx = keep_nonkey | set(best_idx_by_key.values())
    out: list[dict[str, str]] = []
    for idx, row in enumerate(rows):
        if idx in keep_idx and isinstance(row, dict):
            out.append(row)
    return out


def prepare_mapping(
    project_excel: Path,
    source_json_paths: list[Path],
    location: str,
    mapping_json_out: Path,
    mapping_csv_out: Path,
    report_out: Path,
) -> dict[str, Any]:
    wb = load_workbook_safe(project_excel)
    sources = load_source_prices(source_json_paths, location=location)

    mapping_rows: list[dict[str, str]] = []
    used_sources: set[str] = set()

    for sheet in wb.sheetnames:
        if sheet in SKIP_SHEETS:
            mapping_rows.append(
                {
                    "项目文件Sheet": sheet,
                    "最新清单厂家Sheet": "",
                    "状态": "跳过(汇总页)",
                    "说明": "非厂家页",
                }
            )
            continue

        scored: list[tuple[int, SourcePrice]] = []
        for src in sources:
            score = _score_company_match(sheet, src.company)
            if score > 0:
                scored.append((score, src))
        if not scored:
            mapping_rows.append(
                {
                    "项目文件Sheet": sheet,
                    "最新清单厂家Sheet": "",
                    "状态": "未匹配(不更新)",
                    "说明": "图片/文档厂家中无对应项",
                }
            )
            continue

        scored.sort(key=lambda x: x[0], reverse=True)
        top_score = scored[0][0]
        top = [src for score, src in scored if score == top_score]
        if len(top) > 1:
            mapping_rows.append(
                {
                    "项目文件Sheet": sheet,
                    "最新清单厂家Sheet": "",
                    "状态": "待确认(冲突)",
                    "说明": "候选: " + " / ".join(sorted({x.company for x in top})),
                }
            )
            continue

        chosen = top[0]
        used_sources.add(chosen.company)
        mapping_rows.append(
            {
                "项目文件Sheet": sheet,
                "最新清单厂家Sheet": chosen.company,
                "状态": "待确认匹配",
                "说明": "自动建议，需人工确认后改为已确认匹配",
            }
        )

    unmapped_sources = [
        {
            "来源厂家": s.company,
            "来源文件": s.source_file,
            "原因": "在项目报价表中无对应sheet",
        }
        for s in sources
        if s.company not in used_sources
    ]

    _write_mapping_json_csv(mapping_rows, mapping_json_out, mapping_csv_out)

    report = {
        "phase": "prepare",
        "project_excel": project_excel.name,
        "location": location,
        "mapping_json": str(mapping_json_out),
        "mapping_csv": str(mapping_csv_out),
        "total_rows": len(mapping_rows),
        "pending_count": sum(1 for r in mapping_rows if r["状态"] == "待确认匹配"),
        "conflict_count": sum(1 for r in mapping_rows if r["状态"].startswith("待确认(冲突)")),
        "unmatched_count": sum(1 for r in mapping_rows if r["状态"].startswith("未匹配")),
        "unmapped_sources": unmapped_sources,
    }
    report_out.parent.mkdir(parents=True, exist_ok=True)
    report_out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def _source_lookup(sources: list[SourcePrice]) -> dict[str, SourcePrice]:
    by_norm: dict[str, SourcePrice] = {}
    for s in sources:
        key = _normalize_company(s.company)
        if not key:
            continue
        old = by_norm.get(key)
        if old is None or s.quote_date >= old.quote_date:
            by_norm[key] = s
    return by_norm


def apply_writeback(
    project_excel: Path,
    source_json_paths: list[Path],
    mapping_json_path: Path,
    location: str,
    report_out: Path,
    dry_run: bool = False,
) -> dict[str, Any]:
    raw_mapping_rows = json.loads(mapping_json_path.read_text(encoding="utf-8"))
    if not isinstance(raw_mapping_rows, list):
        raise ValueError("mapping json must be a list")
    mapping_rows = _dedupe_mapping_rows(raw_mapping_rows)
    deduped_row_count = max(0, len(raw_mapping_rows) - len(mapping_rows))

    sources = load_source_prices(source_json_paths, location=location)
    source_map = _source_lookup(sources)

    confirmed_or_accepted_sources: set[str] = set()
    for row in mapping_rows:
        if not isinstance(row, dict):
            continue
        status = str(row.get("状态") or "").strip()
        source_company = str(row.get("最新清单厂家Sheet") or "").strip()
        if status in {CONFIRMED_WRITE_STATUS, CONFIRMED_SKIP_STATUS} and source_company:
            norm = _normalize_company(source_company)
            if norm:
                confirmed_or_accepted_sources.add(norm)

    unresolved_sources: list[dict[str, str]] = []
    for src in sources:
        norm = _normalize_company(src.company)
        if norm and norm not in confirmed_or_accepted_sources:
            unresolved_sources.append(
                {
                    "来源厂家": src.company,
                    "来源文件": src.source_file,
                    "原因": "发现新厂家或未确认厂家，请先更新对照表后再继续写价",
                }
            )
    if unresolved_sources:
        report = {
            "phase": "apply",
            "project_excel": project_excel.name,
            "location": location,
            "mapping_json": str(mapping_json_path),
            "mapping_row_count": len(mapping_rows),
            "deduped_row_count": deduped_row_count,
            "blocked": True,
            "blocked_reason": "存在未确认的新厂家，已停止写价",
            "unresolved_sources": unresolved_sources,
            "dry_run": dry_run,
            "updated_count": 0,
            "skipped_count": 0,
            "backup_file": None,
            "updates": [],
            "skipped": [],
        }
        report_out.parent.mkdir(parents=True, exist_ok=True)
        report_out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        return report

    wb = load_workbook_safe(project_excel)
    updates: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for row in mapping_rows:
        if not isinstance(row, dict):
            continue
        sheet = str(row.get("项目文件Sheet") or "").strip()
        source_company = str(row.get("最新清单厂家Sheet") or "").strip()
        status = str(row.get("状态") or "").strip()
        note = str(row.get("说明") or "").strip()

        if status != CONFIRMED_WRITE_STATUS:
            skipped.append(
                {
                    "项目文件Sheet": sheet,
                    "来源厂家": source_company,
                    "原因": f"状态不是{CONFIRMED_WRITE_STATUS}（当前: {status or '空'}）",
                }
            )
            continue
        if not sheet or sheet not in wb.sheetnames:
            skipped.append(
                {"项目文件Sheet": sheet, "来源厂家": source_company, "原因": "项目sheet不存在"}
            )
            continue
        if not source_company:
            skipped.append(
                {"项目文件Sheet": sheet, "来源厂家": source_company, "原因": f"{CONFIRMED_WRITE_STATUS}但来源厂家为空"}
            )
            continue

        src = source_map.get(_normalize_company(source_company))
        if src is None:
            skipped.append(
                {
                    "项目文件Sheet": sheet,
                    "来源厂家": source_company,
                    "原因": "来源厂家在本次图片/文档提取结果中不存在",
                }
            )
            continue
        # Handle electronic negotiation (电议) case
        if src.is_electronic_negotiation:
            skipped.append(
                {
                    "项目文件Sheet": sheet,
                    "来源厂家": source_company,
                    "原因": f"{src.company}报价为电议，本次报价表中{src.company}厂家线下价格未做更新",
                }
            )
            continue

        ws = wb[sheet]
        old_h1, old_h3, old_h4 = ws["H1"].value, ws["H3"].value, ws["H4"].value
        new_h1 = f"报价[{src.quote_date}]"

        deviation_config = PriceDeviationConfig()
        deviation_reasons: list[str] = []
        coil_deviation = _check_price_deviation(
            offline_price=src.coil_price,
            web_price=ws["G3"].value,
            label="盘螺",
            config=deviation_config,
        )
        if coil_deviation:
            deviation_reasons.append(coil_deviation)
        rebar_deviation = _check_price_deviation(
            offline_price=src.rebar_price,
            web_price=ws["G4"].value,
            label="螺纹",
            config=deviation_config,
        )
        if rebar_deviation:
            deviation_reasons.append(rebar_deviation)

        if deviation_reasons:
            skipped.append(
                {
                    "项目文件Sheet": sheet,
                    "来源厂家": source_company,
                    "原因": "；".join(deviation_reasons),
                }
            )
            continue

        reference_notes = _missing_reference_notes(ws, src)
        ws["H1"] = new_h1

        # 只要有一个价格有值就执行回写，缺失的价格保留原值
        has_update = False
        partial_skip = []
        
        if src.coil_price is not None:
            ws["H3"] = src.coil_price
            ws["H3"].font = Font(color="FFFF0000")  # 红色标记已更新
            has_update = True
        else:
            partial_skip.append("盘螺")

        if src.rebar_price is not None:
            ws["H4"] = src.rebar_price
            ws["H4"].font = Font(color="FFFF0000")  # 红色标记已更新
            has_update = True
        else:
            partial_skip.append("螺纹")

        # 如果两个价格都缺失，则跳过
        if not has_update:
            skipped.append(
                {
                    "项目文件Sheet": sheet,
                    "来源厂家": source_company,
                    "原因": "缺少螺纹和盘螺价格",
                }
            )
            continue

        partial_note = note
        if reference_notes:
            ref_note = "；".join(reference_notes)
            partial_note = f"{partial_note}({ref_note})" if partial_note else ref_note
        if partial_skip:
            skip_note = f"跳过{'+'.join(partial_skip)}：价格为空，保留原值"
            if partial_note:
                partial_note = f"{partial_note}({skip_note})"
            else:
                partial_note = skip_note

        updates.append(
            {
                "项目文件Sheet": sheet,
                "来源厂家": source_company,
                "来源文件": src.source_file,
                "备注": partial_note,
                "H1_old": old_h1,
                "H1_new": new_h1,
                "H3_old": old_h3,
                "H3_new": src.coil_price if src.coil_price is not None else f"{old_h3} (保留原值)",
                "H4_old": old_h4,
                "H4_new": src.rebar_price if src.rebar_price is not None else f"{old_h4} (保留原值)",
                "partial_update": bool(partial_skip),
                "skipped_items": partial_skip if partial_skip else None,
            }
        )

    if dry_run:
        backup_file = None
        wb.close()
    elif updates:
        backup_dir = project_excel.parent / "备份"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = backup_dir / f"{project_excel.stem}.backup_before_image_doc_write_{stamp}.xlsx"
        shutil.copy2(project_excel, backup_file)
        wb.save(project_excel)
        wb.close()
    else:
        backup_file = None
        wb.close()

    inventory_review = build_inventory_review(source_json_paths)

    if dry_run:
        inventory_report = {
            "status": "review_only",
            "reason": "dry-run模式不修改库存颜色",
            "review": inventory_review,
        }
    else:
        inventory_report = None
        try:
            mill_inventories = inventory_items_from_review(inventory_review)
            if mill_inventories:
                inventory_report = apply_inventory_to_project(
                    project_excel=project_excel,
                    mill_inventories=mill_inventories,
                    sheet_name="报价表",
                    mapping_json_path=mapping_json_path,
                    clear_existing_colors=True,
                )
        except Exception as exc:
            inventory_report = {"status": "error", "error": str(exc)}
        if inventory_report is not None:
            inventory_report["review"] = inventory_review

    report = {
        "phase": "apply",
        "project_excel": project_excel.name,
        "location": location,
        "mapping_json": str(mapping_json_path),
        "mapping_row_count": len(mapping_rows),
        "deduped_row_count": deduped_row_count,
        "dry_run": dry_run,
        "updated_count": len(updates),
        "skipped_count": len(skipped),
        "backup_file": str(backup_file) if backup_file else None,
        "updates": updates,
        "skipped": skipped,
        "inventory_report": inventory_report,
    }
    report_out.parent.mkdir(parents=True, exist_ok=True)
    report_out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Image/doc price writeback in two phases: prepare mapping then apply confirmed mapping."
    )
    sub = p.add_subparsers(dest="command", required=True)

    p_prepare = sub.add_parser("prepare", help="Generate pending mapping only (no writeback).")
    p_prepare.add_argument("--project", required=True, help="Project excel path")
    p_prepare.add_argument("--location", required=True, help="Target location, e.g. 蚌埠")
    p_prepare.add_argument("--sources", nargs="+", required=True, help="OCR extraction json files")
    p_prepare.add_argument("--mapping-json", help="Output mapping json path")
    p_prepare.add_argument("--mapping-csv", help="Output mapping csv path")
    p_prepare.add_argument("--report-out", help="Prepare report output path")

    p_apply = sub.add_parser("apply", help="Apply writeback using confirmed mapping.")
    p_apply.add_argument("--project", required=True, help="Project excel path")
    p_apply.add_argument("--location", required=True, help="Target location, e.g. 蚌埠")
    p_apply.add_argument("--sources", nargs="+", required=True, help="OCR extraction json files")
    p_apply.add_argument("--mapping-json", required=True, help="Confirmed mapping json path")
    p_apply.add_argument("--report-out", help="Apply report output path")
    return p


def main() -> int:
    args = _build_parser().parse_args()
    project = Path(args.project)
    source_paths = [Path(x) for x in args.sources]
    location = args.location
    city_tag = "安徽蚌埠" if "蚌埠" in location else location

    if args.command == "prepare":
        default_prefix = Path("运行产物") / f"图片文档厂家对照表_{city_tag}_待确认"
        mapping_json = Path(args.mapping_json) if args.mapping_json else default_prefix.with_suffix(".json")
        mapping_csv = Path(args.mapping_csv) if args.mapping_csv else default_prefix.with_suffix(".csv")
        report_out = (
            Path(args.report_out)
            if args.report_out
            else Path("运行产物") / f"图片文档回写准备报告_{city_tag}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.json"
        )
        report = prepare_mapping(
            project_excel=project,
            source_json_paths=source_paths,
            location=location,
            mapping_json_out=mapping_json,
            mapping_csv_out=mapping_csv,
            report_out=report_out,
        )
        print(f"Mapping JSON: {report['mapping_json']}")
        print(f"Mapping CSV: {report['mapping_csv']}")
        print(f"Pending: {report['pending_count']}")
        print(f"Conflicts: {report['conflict_count']}")
        print(f"Unmatched: {report['unmatched_count']}")
        print(f"Report: {report_out}")
        print("Next: 请人工确认对照表，将状态改为“已确认匹配”或“已确认不更新”后再执行 apply。")
        return 0

    report_out = (
        Path(args.report_out)
        if args.report_out
        else Path("运行产物") / f"图片文档回写报告_{city_tag}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.json"
    )
    report = apply_writeback(
        project_excel=project,
        source_json_paths=source_paths,
        mapping_json_path=Path(args.mapping_json),
        location=location,
        report_out=report_out,
    )
    if report.get("blocked"):
        print("Blocked: 存在未确认的新厂家，未执行写价。")
        for item in report.get("unresolved_sources", []):
            print(f"- {item.get('来源厂家')} ({item.get('来源文件')}): {item.get('原因')}")
    else:
        print(f"Updated: {report['updated_count']}")
        print(f"Skipped: {report['skipped_count']}")
        print(f"Backup: {report['backup_file']}")
    print(f"Report: {report_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
