from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
from contextlib import contextmanager
from datetime import date, datetime
from pathlib import Path
from typing import Any

from .web_price import (
    CONFIRMED_SKIP_STATUS as WEB_CONFIRMED_SKIP_STATUS,
    CONFIRMED_WRITE_STATUS as WEB_CONFIRMED_WRITE_STATUS,
    _city,
    _norm_company,
    _parse_location,
    apply_web_writeback,
    fetch_web_prices,
    prepare_web_mapping,
)
from .writeback_image_doc import (
    CONFIRMED_SKIP_STATUS as IMG_CONFIRMED_SKIP_STATUS,
    CONFIRMED_WRITE_STATUS as IMG_CONFIRMED_WRITE_STATUS,
    apply_writeback as apply_image_writeback,
    load_source_prices,
    prepare_mapping as prepare_image_mapping,
)
from .reporting import render_single_report_markdown
from .audit import audit_image_doc_updates
from .xlsx_utils import load_workbook_safe

SKIP_SHEETS = {"报价表"}
DEFAULT_OFFLINE_DIR = Path("线下报价")
OFFLINE_SOURCE_SUFFIXES = {".jpg", ".jpeg", ".png", ".pdf", ".txt", ".json"}
OCR_JSON_GLOB = "ocr价格提取_*.json"


def _ts() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H%M%S")


def _safe_locations(project: Path) -> tuple[str, str]:
    try:
        return _parse_location(project.name)
    except Exception:
        return "未知地点", "未知地点"


def _json_read(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _json_write(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
            default=lambda obj: obj.isoformat() if isinstance(obj, (datetime, date)) else str(obj),
        ),
        encoding="utf-8",
    )


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _assert_dry_run_unchanged(project: Path, before_hash: str) -> None:
    after_hash = _file_sha256(project)
    if after_hash != before_hash:
        raise RuntimeError(
            f"dry-run 修改了项目 Excel，禁止继续：{project} before={before_hash} after={after_hash}"
        )


@contextmanager
def _pipeline_lock(lock_path: Path, project: Path):
    if lock_path.exists():
        raise RuntimeError(f"已有报价更新任务锁，请先人工确认并清理：{lock_path}")
    payload = {
        "pid": os.getpid(),
        "project": str(project),
        "started_at": datetime.now().isoformat(timespec="seconds"),
    }
    _json_write(lock_path, payload)
    try:
        yield
    finally:
        try:
            lock_path.unlink()
        except FileNotFoundError:
            pass


def _manifest_path(artifact_dir: Path) -> Path:
    return artifact_dir / "dry_run_manifest.json"


def _write_manifest(path: Path, result: dict[str, Any], artifact_dir: Path) -> None:
    manifest = {
        "project": result.get("project"),
        "artifact_dir": str(artifact_dir),
        "mode": result.get("mode"),
        "started_at": result.get("started_at"),
        "ended_at": result.get("ended_at"),
        "project_excel_hash_before": result.get("project_excel_hash_before"),
        "project_excel_hash_after": result.get("project_excel_hash_after"),
        "web": result.get("web"),
        "image_doc": result.get("image_doc"),
    }
    _json_write(path, manifest)


def _read_manifest(path: Path) -> dict[str, Any]:
    payload = _json_read(path)
    if not isinstance(payload, dict):
        raise RuntimeError(f"manifest不是JSON对象：{path}")
    return payload


def _web_apply_from_manifest(project: Path, manifest: dict[str, Any]) -> dict[str, Any]:
    artifact_dir = Path(str(manifest.get("artifact_dir") or ""))
    web = manifest.get("web") if isinstance(manifest.get("web"), dict) else {}
    location = str(web.get("location") or manifest.get("web_location") or "")
    if not artifact_dir.exists():
        raise RuntimeError(f"manifest产物目录不存在：{artifact_dir}")
    return _web_apply_from_artifacts(project=project, location=location, artifact_dir=artifact_dir)


def _image_apply_from_manifest(project: Path, manifest: dict[str, Any]) -> dict[str, Any]:
    artifact_dir = Path(str(manifest.get("artifact_dir") or ""))
    image_doc = manifest.get("image_doc") if isinstance(manifest.get("image_doc"), dict) else {}
    location = str(image_doc.get("location") or manifest.get("image_location") or "")
    source_jsons = [Path(x) for x in image_doc.get("source_jsons", []) or []]
    return _image_apply_from_artifacts(project=project, location=location, artifact_dir=artifact_dir, source_jsons=source_jsons)


def _mapping_has_pending(path: Path) -> bool:
    if not path.exists():
        return False
    payload = _json_read(path)
    if not isinstance(payload, list):
        return False
    for row in payload:
        if not isinstance(row, dict):
            continue
        if str(row.get("状态") or "").strip().startswith("待确认"):
            return True
    return False


def _confirmed_sources(mapping_rows: list[dict[str, Any]], write_status: str, skip_status: str) -> set[str]:
    out: set[str] = set()
    for row in mapping_rows:
        if not isinstance(row, dict):
            continue
        status = str(row.get("状态") or "").strip()
        source = str(row.get("最新清单厂家Sheet") or "").strip()
        if status in {write_status, skip_status} and source:
            out.add(_norm_company(source))
    return out


def _mapping_project_set(mapping_rows: list[dict[str, Any]]) -> set[str]:
    out: set[str] = set()
    for row in mapping_rows:
        if not isinstance(row, dict):
            continue
        sheet = str(row.get("项目文件Sheet") or "").strip()
        if sheet and sheet not in SKIP_SHEETS:
            out.add(sheet)
    return out


def _project_sheet_set(project_excel: Path) -> set[str]:
    wb = load_workbook_safe(project_excel, data_only=True)
    return {x for x in wb.sheetnames if x not in SKIP_SHEETS}


def _safe_project_dir_name(project: Path) -> str:
    name = project.stem.strip()
    if not name:
        name = "未命名项目"
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    return name.rstrip(". ") or "未命名项目"


def _project_artifact_dir(base_dir: Path, project: Path) -> Path:
    return base_dir / _safe_project_dir_name(project)


def _default_ocr_output(input_file: Path, artifact_dir: Path) -> Path:
    return artifact_dir / f"ocr价格提取_{input_file.stem}.json"


def _run_ocr_for_inputs(inputs: list[Path], location: str, artifact_dir: Path) -> list[Path]:
    """Run OCR for input files. All image/PDF files use MiniMax vision recognition."""
    outputs: list[Path] = []
    # Set UTF-8 encoding for child processes to avoid UnicodeEncodeError on Windows
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    for src in inputs:
        out = _default_ocr_output(src, artifact_dir)
        cmd = [
            sys.executable,
            "-m",
            "ocr_price.cli",
            "--input",
            str(src),
            "--location",
            location,
            "--output",
            str(out),
        ]
        # All image and PDF files use MiniMax vision recognition by default
        # Text files (.txt) will be auto-detected and processed with text parser
        subprocess.run(cmd, check=True, env=env)
        outputs.append(out)
    return outputs


def _pending_details(mapping_json: Path, project_excel: Path | None = None) -> dict[str, Any]:
    """获取待确认详情，包括Excel全部厂家清单"""
    result = {
        "all_factories": [],
        "pending_matches": [],
        "pending_new": [],
        "unmatched": [],
    }
    
    # 读取Excel全部厂家
    if project_excel and project_excel.exists():
        try:
            wb = load_workbook_safe(project_excel, data_only=True)
            for sheet in wb.sheetnames:
                if sheet not in SKIP_SHEETS:
                    result["all_factories"].append(sheet)
        except Exception:
            pass
    
    if not mapping_json.exists():
        return result
    
    payload = _json_read(mapping_json)
    if not isinstance(payload, list):
        return result
    
    for row in payload:
        if not isinstance(row, dict):
            continue
        status = str(row.get("状态") or "").strip()
        sheet = str(row.get("项目文件Sheet") or "").strip()
        source = str(row.get("最新清单厂家Sheet") or "").strip()
        note = str(row.get("说明") or "").strip()
        
        item = {
            "项目文件Sheet": sheet,
            "最新清单厂家Sheet": source,
            "状态": status,
            "说明": note,
        }
        
        if status.startswith("待确认匹配"):
            result["pending_matches"].append(item)
        elif status.startswith("待确认(新厂家)"):
            result["pending_new"].append(item)
        elif status.startswith("未匹配"):
            result["unmatched"].append(item)
    
    return result


def _web_apply_summary(apply_report: dict[str, Any]) -> dict[str, Any]:
    updated_items: list[dict[str, Any]] = []
    for row in apply_report.get("updated_sheets", []) or []:
        if not isinstance(row, dict):
            continue
        updated_items.append(
            {
                "项目文件Sheet": str(row.get("sheet") or ""),
                "来源厂家": str(row.get("mill") or ""),
                "G1": {"old": row.get("G1_old"), "new": row.get("G1_new")},
                "G3": {"old": row.get("G3_old"), "new": row.get("G3_coil")},
                "G4": {"old": row.get("G4_old"), "new": row.get("G4_rebar")},
            }
        )
    skipped_items: list[dict[str, Any]] = []
    for row in apply_report.get("skipped", []) or []:
        if not isinstance(row, dict):
            continue
        skipped_items.append(
            {
                "项目文件Sheet": str(row.get("sheet") or ""),
                "来源厂家": str(row.get("mill") or ""),
                "原因": str(row.get("reason") or ""),
            }
        )
    return {
        "updated_count": int(apply_report.get("updated_sheet_count") or len(updated_items)),
        "skipped_count": int(apply_report.get("skipped_count") or len(skipped_items)),
        "updated_items": updated_items,
        "skipped_items": skipped_items,
    }


def _image_apply_summary(apply_report: dict[str, Any]) -> dict[str, Any]:
    updated_items: list[dict[str, Any]] = []
    for row in apply_report.get("updates", []) or []:
        if not isinstance(row, dict):
            continue
        updated_items.append(
            {
                "项目文件Sheet": str(row.get("项目文件Sheet") or ""),
                "来源厂家": str(row.get("来源厂家") or ""),
                "H1": {"old": row.get("H1_old"), "new": row.get("H1_new")},
                "H3": {"old": row.get("H3_old"), "new": row.get("H3_new")},
                "H4": {"old": row.get("H4_old"), "new": row.get("H4_new")},
            }
        )
    skipped_items: list[dict[str, Any]] = []
    for row in apply_report.get("skipped", []) or []:
        if not isinstance(row, dict):
            continue
        skipped_items.append(
            {
                "项目文件Sheet": str(row.get("项目文件Sheet") or ""),
                "来源厂家": str(row.get("来源厂家") or ""),
                "原因": str(row.get("原因") or ""),
            }
        )
    return {
        "updated_count": int(apply_report.get("updated_count") or len(updated_items)),
        "skipped_count": int(apply_report.get("skipped_count") or len(skipped_items)),
        "updated_items": updated_items,
        "skipped_items": skipped_items,
    }


def _format_price_change(old_val: Any, new_val: Any) -> str:
    """格式化价格变动：老值 → 新值"""
    old_str = str(old_val) if old_val is not None else "null"
    new_str = str(new_val) if new_val is not None else "null"
    return f"{old_str} → {new_str}"


def _generate_web_report(summary: dict[str, Any]) -> str:
    """生成网价更新报告（符合commands.md模板）"""
    lines: list[str] = []
    updated = summary.get("updated_items", []) or []
    skipped = summary.get("skipped_items", []) or []
    
    # 分类：完整更新和部分更新
    full_updates = [x for x in updated if not x.get("partial_update")]
    partial_updates = [x for x in updated if x.get("partial_update")]
    
    lines.append(f"## 网价更新明细（{len(full_updates)}个厂家）")
    lines.append("| 序号 | 厂家 | G3(盘螺) 老→新 | G4(螺纹) 老→新 |")
    lines.append("|------|------|----------------|----------------|")
    for idx, row in enumerate(full_updates, 1):
        sheet = str(row.get("项目文件Sheet") or "")
        g3_old = row.get("G3", {}).get("old")
        g3_new = row.get("G3", {}).get("new")
        g4_old = row.get("G4", {}).get("old")
        g4_new = row.get("G4", {}).get("new")
        lines.append(f"| {idx} | {sheet} | {_format_price_change(g3_old, g3_new)} | {_format_price_change(g4_old, g4_new)} |")
    
    if partial_updates:
        lines.append(f"\n## 网价部分更新明细（{len(partial_updates)}个厂家）")
        lines.append("| 序号 | 厂家 | G3(盘螺) 老→新 | G4(螺纹) 老→新 | 说明 |")
        lines.append("|------|------|----------------|----------------|------|")
        for idx, row in enumerate(partial_updates, 1):
            sheet = str(row.get("项目文件Sheet") or "")
            g3_old = row.get("G3", {}).get("old")
            g3_new = row.get("G3", {}).get("new")
            g4_old = row.get("G4", {}).get("old")
            g4_new = row.get("G4", {}).get("new")
            skipped_items = row.get("skipped_items", [])
            note = f"跳过{'+'.join(skipped_items)}：价格为空，保留原值" if skipped_items else ""
            lines.append(f"| {idx} | {sheet} | {_format_price_change(g3_old, g3_new)} | {_format_price_change(g4_old, g4_new)} | {note} |")
    
    if skipped:
        lines.append(f"\n**跳过（{len(skipped)}个）：**")
        for row in skipped:
            sheet = str(row.get("项目文件Sheet") or "")
            reason = str(row.get("原因") or "")
            lines.append(f"- {sheet} - {reason}")
    
    return "\n".join(lines)


def _generate_image_report(summary: dict[str, Any]) -> str:
    """生成图片/文档更新报告（符合commands.md模板）"""
    lines: list[str] = []
    updated = summary.get("updated_items", []) or []
    skipped = summary.get("skipped_items", []) or []
    
    # 分类：完整更新和部分更新
    full_updates = [x for x in updated if not x.get("partial_update")]
    partial_updates = [x for x in updated if x.get("partial_update")]
    
    lines.append(f"## 图片/文档更新明细（{len(full_updates)}个厂家）")
    lines.append("| 序号 | 厂家 | H3(盘螺) 老→新 | H4(螺纹) 老→新 |")
    lines.append("|------|------|----------------|----------------|")
    for idx, row in enumerate(full_updates, 1):
        sheet = str(row.get("项目文件Sheet") or "")
        h3_old = row.get("H3", {}).get("old")
        h3_new = row.get("H3", {}).get("new")
        h4_old = row.get("H4", {}).get("old")
        h4_new = row.get("H4", {}).get("new")
        lines.append(f"| {idx} | {sheet} | {_format_price_change(h3_old, h3_new)} | {_format_price_change(h4_old, h4_new)} |")
    
    if partial_updates:
        lines.append(f"\n## 图片/文档部分更新明细（{len(partial_updates)}个厂家）")
        lines.append("| 序号 | 厂家 | H3(盘螺) 老→新 | H4(螺纹) 老→新 | 说明 |")
        lines.append("|------|------|----------------|----------------|------|")
        for idx, row in enumerate(partial_updates, 1):
            sheet = str(row.get("项目文件Sheet") or "")
            h3_old = row.get("H3", {}).get("old")
            h3_new = row.get("H3", {}).get("new")
            h4_old = row.get("H4", {}).get("old")
            h4_new = row.get("H4", {}).get("new")
            skipped_items = row.get("skipped_items", [])
            note = f"跳过{'+'.join(skipped_items)}：价格为空，保留原值" if skipped_items else ""
            lines.append(f"| {idx} | {sheet} | {_format_price_change(h3_old, h3_new)} | {_format_price_change(h4_old, h4_new)} | {note} |")
    
    if skipped:
        lines.append(f"\n**跳过（{len(skipped)}个）：**")
        for row in skipped:
            sheet = str(row.get("项目文件Sheet") or "")
            reason = str(row.get("原因") or "")
            lines.append(f"- {sheet} - {reason}")
    
    return "\n".join(lines)


def _generate_inventory_report(inventory_report: dict[str, Any] | None) -> str:
    """生成库存颜色标注报告"""
    if not inventory_report:
        return ""
    
    lines: list[str] = []
    lines.append("\n**【新增】库存颜色标注：**")
    
    # 统计各状态数量
    status_counts = {"充足": 0, "告警": 0, "缺货": 0}
    factory_details = []
    
    for file_info in inventory_report.get("factory_details", []) or []:
        for detail in file_info.get("details", []) or []:
            status = detail.get("status", "")
            if status in status_counts:
                status_counts[status] += 1
    
    lines.append(f"- 蓝色（充足）：{status_counts['充足']}个")
    lines.append(f"- 黄色（告警）：{status_counts['告警']}个")
    lines.append(f"- 红色（缺货）：{status_counts['缺货']}个")
    
    # 涉及厂家
    factories = set()
    for file_info in inventory_report.get("factory_details", []) or []:
        file_name = file_info.get("file", "")
        # 从文件名提取厂家名
        factory = file_name.replace("ocr价格提取_", "").replace(".json", "")
        factories.add(factory)
    
    if factories:
        lines.append(f"- 涉及厂家：{', '.join(sorted(factories))}")
    
    return "\n".join(lines)


def _format_change(cell: Any) -> str:
    if not isinstance(cell, dict):
        return "空 → 空"
    old = cell.get("old")
    new = cell.get("new")
    old_text = str(old) if old is not None else "空"
    new_text = str(new) if new is not None else "空"
    return f"{old_text} → {new_text}"


def _print_flow_details(flow_name: str, flow: Any) -> None:
    if not isinstance(flow, dict):
        return
    status = str(flow.get("status") or "")
    if status == "pending_confirmation":
        pending = flow.get("pending_details") or []
        print(f"{flow_name} Pending: {len(pending)}")
        for row in pending:
            print(
                f"{flow_name} PendingItem: "
                f"{json.dumps(row, ensure_ascii=False)}"
            )
        return
    if status != "ok":
        return
    summary = flow.get("apply_summary") or {}
    print(
        f"{flow_name} Updated={summary.get('updated_count', 0)} "
        f"Skipped={summary.get('skipped_count', 0)}"
    )
    for row in summary.get("updated_items", []) or []:
        if flow_name == "Web":
            line = (
                f"{row.get('项目文件Sheet', '')} | {row.get('来源厂家', '')} | "
                f"G1: {_format_change(row.get('G1'))} | "
                f"G3: {_format_change(row.get('G3'))} | "
                f"G4: {_format_change(row.get('G4'))}"
            )
        else:
            line = (
                f"{row.get('项目文件Sheet', '')} | {row.get('来源厂家', '')} | "
                f"H1: {_format_change(row.get('H1'))} | "
                f"H3: {_format_change(row.get('H3'))} | "
                f"H4: {_format_change(row.get('H4'))}"
            )
        print(f"{flow_name} UpdatedItem: {line}")
    for row in summary.get("skipped_items", []) or []:
        print(f"{flow_name} SkippedItem: {json.dumps(row, ensure_ascii=False)}")


def _print_result_details(result: dict[str, Any]) -> None:
    project_name = Path(str(result.get("project") or "")).name or str(result.get("project") or "")
    print(f"Project: {project_name}")
    print(f"ResultStatus: {result.get('status')}")
    _print_flow_details("Web", result.get("web"))
    _print_flow_details("ImageDoc", result.get("image_doc"))


def _latest_file(paths: list[Path]) -> Path | None:
    existing = [p for p in paths if p.exists() and p.is_file()]
    if not existing:
        return None
    return max(existing, key=lambda p: p.stat().st_mtime)


def _latest_glob_file(base_dir: Path, pattern: str) -> Path | None:
    return _latest_file([p for p in base_dir.glob(pattern) if p.is_file()])


def _artifact_mapping_path(artifact_dir: Path, stem: str) -> Path | None:
    confirmed = artifact_dir / f"{stem}_已确认.json"
    pending = artifact_dir / f"{stem}_待确认.json"
    if confirmed.exists():
        return confirmed
    if pending.exists():
        return pending
    return None


def _web_apply_from_artifacts(
    project: Path,
    location: str,
    artifact_dir: Path,
) -> dict[str, Any]:
    mapping_path = _artifact_mapping_path(artifact_dir, f"厂家对照表_{location}")
    if mapping_path is None:
        return {
            "status": "skipped",
            "phase": "web_apply",
            "reason": "未找到网价对照表产物，请先执行 --dry-run 或加 --refresh-web-artifacts。",
        }
    if _mapping_has_pending(mapping_path):
        return {
            "status": "pending_confirmation",
            "phase": "web_apply",
            "pending_mapping_json": str(mapping_path),
            "pending_details": _pending_details(mapping_path, project),
            "reason": "网价对照存在待确认项，请先确认后再执行写入",
        }

    fetch_report_path = _latest_glob_file(artifact_dir, f"网价提取报告_{location}_*.json")
    fetch_report: dict[str, Any] = {}
    if fetch_report_path is not None:
        payload = _json_read(fetch_report_path)
        if isinstance(payload, dict):
            fetch_report = payload

    raw_excel: Path | None = None
    raw_from_report = str(fetch_report.get("raw_price_excel") or "").strip()
    if raw_from_report:
        candidate = Path(raw_from_report)
        if not candidate.is_absolute():
            candidate = artifact_dir / candidate
        if candidate.exists():
            raw_excel = candidate
    if raw_excel is None:
        raw_excel = _latest_glob_file(artifact_dir, f"{location}*建筑钢材原料价格清单*.xlsx")
    if raw_excel is None:
        raw_excel = _latest_glob_file(artifact_dir, "*建筑钢材原料价格清单*.xlsx")
    if raw_excel is None:
        return {
            "status": "skipped",
            "phase": "web_apply",
            "reason": "未找到网价清单产物，请先执行 --dry-run 或加 --refresh-web-artifacts。",
        }

    quote_date = str(fetch_report.get("quote_date") or "").strip() or datetime.now().strftime("%Y-%m-%d")
    source_url = str(fetch_report.get("latest_url") or "").strip()
    apply_report_path = artifact_dir / f"网价回写报告_{location}_{_ts()}.json"
    apply_report = apply_web_writeback(
        project_excel=project,
        raw_excel=raw_excel,
        mapping_json=mapping_path,
        quote_date=quote_date,
        source_url=source_url,
        report_out=apply_report_path,
    )
    if apply_report.get("blocked"):
        return {
            "status": "pending_confirmation",
            "phase": "web_apply",
            "apply_report": str(apply_report_path),
            "reason": str(apply_report.get("blocked_reason") or "网价回写被阻断"),
        }
    return {
        "status": "ok",
        "phase": "web_apply",
        "apply_report": str(apply_report_path),
        "source_mode": "reuse_artifacts",
        "apply_summary": _web_apply_summary(apply_report),
    }


def _image_apply_from_artifacts(
    project: Path,
    location: str,
    artifact_dir: Path,
    source_jsons: list[Path] | None = None,
) -> dict[str, Any]:
    mapping_path = _artifact_mapping_path(artifact_dir, f"图片文档厂家对照表_{location}")
    if mapping_path is None:
        return {
            "status": "skipped",
            "phase": "image_apply",
            "reason": "未找到图片/文档对照表产物，请先执行 --dry-run 或加 --refresh-image-artifacts。",
        }
    if _mapping_has_pending(mapping_path):
        return {
            "status": "pending_confirmation",
            "phase": "image_apply",
            "pending_mapping_json": str(mapping_path),
            "pending_details": _pending_details(mapping_path, project),
            "reason": "图片/文档对照存在待确认项，请先确认后再执行写入",
        }

    json_paths = [p for p in (source_jsons or []) if p.exists() and p.is_file()]
    if not json_paths:
        json_paths = sorted([p for p in artifact_dir.glob(OCR_JSON_GLOB) if p.is_file()])
    if not json_paths:
        return {
            "status": "skipped",
            "phase": "image_apply",
            "reason": "未找到OCR JSON产物，请先执行 --dry-run 或加 --refresh-image-artifacts。",
        }

    apply_report_path = artifact_dir / f"图片文档回写报告_{location}_{_ts()}.json"
    apply_report = apply_image_writeback(
        project_excel=project,
        source_json_paths=json_paths,
        mapping_json_path=mapping_path,
        location=_city(location),
        report_out=apply_report_path,
        dry_run=False,
    )
    if apply_report.get("blocked"):
        return {
            "status": "pending_confirmation",
            "phase": "image_apply",
            "apply_report": str(apply_report_path),
            "reason": str(apply_report.get("blocked_reason") or "图片/文档回写被阻断"),
        }
    apply_summary = _image_apply_summary(apply_report)
    audit_report = None
    if apply_summary.get("updated_items"):
        audit_report = audit_image_doc_updates(project, apply_report.get("updates") or [])
    inventory_report = apply_report.get("inventory_report")
    return {
        "status": "ok",
        "phase": "image_apply",
        "apply_report": str(apply_report_path),
        "source_mode": "reuse_artifacts",
        "apply_summary": apply_summary,
        "audit_report": audit_report,
        "inventory_report": inventory_report,
    }


def _web_flow(
    project: Path,
    location: str,
    list_url: str | None,
    detail_url: str | None,
    account_file: Path,
    username: str | None,
    password: str | None,
    headless: bool,
    artifact_dir: Path,
    apply_if_ready: bool = True,
    manual_login_timeout_seconds: int = 180,
) -> dict[str, Any]:
    fetch_report_path = artifact_dir / f"网价提取报告_{location}_{_ts()}.json"
    raw_tmp = artifact_dir / f"_tmp_{location}_网价清单_{_ts()}.xlsx"
    fetch_report = fetch_web_prices(
        project_excel=project,
        location=location,
        list_url=list_url,
        detail_url=detail_url,
        account_file=account_file,
        username=username,
        password=password,
        output_excel=raw_tmp,
        report_out=fetch_report_path,
        headless=headless,
        manual_login_timeout_seconds=manual_login_timeout_seconds,
    )
    raw_excel = artifact_dir / f"{location}{fetch_report['quote_date']}建筑钢材原料价格清单.xlsx"
    if raw_excel.exists():
        raw_excel = artifact_dir / f"{location}{fetch_report['quote_date']}建筑钢材原料价格清单_{_ts()}.xlsx"
    raw_excel.parent.mkdir(parents=True, exist_ok=True)
    if raw_tmp.exists():
        shutil.move(str(raw_tmp), str(raw_excel))
    fetch_report["raw_price_excel"] = str(raw_excel)
    _json_write(fetch_report_path, fetch_report)

    prepare_report_path = artifact_dir / f"网价回写准备报告_{location}_{_ts()}.json"
    pending_json = artifact_dir / f"厂家对照表_{location}_待确认.json"
    pending_csv = artifact_dir / f"厂家对照表_{location}_待确认.csv"
    official_json = artifact_dir / f"厂家对照表_{location}_正式.json"
    official_csv = artifact_dir / f"厂家对照表_{location}_正式.csv"

    # 若用户已把待确认表全部改成“已确认*”，则自动提升为正式对照后继续。
    if pending_json.exists() and not _mapping_has_pending(pending_json):
        shutil.copy2(pending_json, official_json)
        if pending_csv.exists():
            shutil.copy2(pending_csv, official_csv)

    prepare_report = prepare_web_mapping(
        project_excel=project,
        raw_excel=raw_excel,
        official_mapping_json=official_json,
        pending_mapping_json=pending_json,
        pending_mapping_csv=pending_csv,
        report_out=prepare_report_path,
    )
    if prepare_report.get("needs_user_confirmation"):
        pending_details = _pending_details(pending_json, project)
        return {
            "status": "pending_confirmation",
            "phase": "web_prepare",
            "fetch_report": str(fetch_report_path),
            "prepare_report": str(prepare_report_path),
            "pending_mapping_json": str(pending_json),
            "pending_mapping_csv": str(pending_csv),
            "pending_details": pending_details,
            "reason": "网价对照存在待确认项，请先确认后再执行",
        }

    shutil.copy2(pending_json, official_json)
    if pending_csv.exists():
        shutil.copy2(pending_csv, official_csv)

    if not apply_if_ready:
        return {
            "status": "prepared",
            "phase": "web_prepare",
            "fetch_report": str(fetch_report_path),
            "prepare_report": str(prepare_report_path),
            "reason": "dry-run模式不写入项目Excel",
        }

    apply_report_path = artifact_dir / f"网价回写报告_{location}_{_ts()}.json"
    apply_report = apply_web_writeback(
        project_excel=project,
        raw_excel=raw_excel,
        mapping_json=official_json,
        quote_date=str(fetch_report.get("quote_date") or datetime.now().strftime("%Y-%m-%d")),
        source_url=str(fetch_report.get("latest_url") or ""),
        report_out=apply_report_path,
    )
    if apply_report.get("blocked"):
        return {
            "status": "pending_confirmation",
            "phase": "web_apply",
            "fetch_report": str(fetch_report_path),
            "prepare_report": str(prepare_report_path),
            "apply_report": str(apply_report_path),
            "reason": str(apply_report.get("blocked_reason") or "网价回写被阻断"),
        }
    apply_summary = _web_apply_summary(apply_report)
    return {
        "status": "ok",
        "phase": "web_apply",
        "fetch_report": str(fetch_report_path),
        "prepare_report": str(prepare_report_path),
        "apply_report": str(apply_report_path),
        "apply_summary": apply_summary,
    }


def _image_flow(
    project: Path,
    location: str,
    source_jsons: list[Path],
    artifact_dir: Path,
    apply_if_ready: bool = True,
    dry_run: bool = False,
) -> dict[str, Any]:
    if not source_jsons:
        return {"status": "skipped", "reason": "未提供图片/文档来源json"}

    confirmed_json = artifact_dir / f"图片文档厂家对照表_{location}_已确认.json"
    confirmed_csv = artifact_dir / f"图片文档厂家对照表_{location}_已确认.csv"
    pending_json = artifact_dir / f"图片文档厂家对照表_{location}_待确认.json"
    pending_csv = artifact_dir / f"图片文档厂家对照表_{location}_待确认.csv"
    prepare_report_path = artifact_dir / f"图片文档回写准备报告_{location}_{_ts()}.json"

    if pending_json.exists() and not _mapping_has_pending(pending_json):
        shutil.copy2(pending_json, confirmed_json)
        if pending_csv.exists():
            shutil.copy2(pending_csv, confirmed_csv)

    source_prices = load_source_prices(source_jsons, location=_city(location))
    source_set = {_norm_company(x.company) for x in source_prices if _norm_company(x.company)}

    reusable = False
    if confirmed_json.exists():
        rows = _json_read(confirmed_json)
        if isinstance(rows, list):
            confirmed_set = _confirmed_sources(rows, IMG_CONFIRMED_WRITE_STATUS, IMG_CONFIRMED_SKIP_STATUS)
            mapping_project_set = _mapping_project_set(rows)
            project_set = _project_sheet_set(project)
            reusable = confirmed_set == source_set and mapping_project_set == project_set

    if not reusable:
        prepare_report = prepare_image_mapping(
            project_excel=project,
            source_json_paths=source_jsons,
            location=_city(location),
            mapping_json_out=pending_json,
            mapping_csv_out=pending_csv,
            report_out=prepare_report_path,
        )
        if prepare_report.get("pending_count", 0) > 0 or prepare_report.get("conflict_count", 0) > 0:
            pending_details = _pending_details(pending_json, project)
            return {
                "status": "pending_confirmation",
                "phase": "image_prepare",
                "prepare_report": str(prepare_report_path),
                "pending_mapping_json": str(pending_json),
                "pending_mapping_csv": str(pending_csv),
                "pending_details": pending_details,
                "reason": "图片/文档对照存在待确认项，请先确认后再执行",
            }
        shutil.copy2(pending_json, confirmed_json)
        if pending_csv.exists():
            shutil.copy2(pending_csv, confirmed_csv)

    if not apply_if_ready:
        return {
            "status": "prepared",
            "phase": "image_prepare",
            "prepare_report": str(prepare_report_path),
            "reused_confirmed_mapping": reusable,
            "reason": "已生成图片/文档厂家对照，等待确认后再回写",
        }

    apply_report_path = artifact_dir / f"图片文档回写报告_{location}_{_ts()}.json"
    apply_report = apply_image_writeback(
        project_excel=project,
        source_json_paths=source_jsons,
        mapping_json_path=confirmed_json if reusable or confirmed_json.exists() else pending_json,
        location=_city(location),
        report_out=apply_report_path,
        dry_run=dry_run,
    )
    if apply_report.get("blocked"):
        return {
            "status": "pending_confirmation",
            "phase": "image_apply",
            "apply_report": str(apply_report_path),
            "reason": str(apply_report.get("blocked_reason") or "图片/文档回写被阻断"),
        }
    apply_summary = _image_apply_summary(apply_report)
    audit_report = None
    if not apply_report.get("dry_run") and apply_summary.get("updated_items"):
        audit_report = audit_image_doc_updates(project, apply_report.get("updates") or [])
    return {
        "status": "ok",
        "phase": "image_apply",
        "apply_report": str(apply_report_path),
        "reused_confirmed_mapping": reusable,
        "apply_summary": apply_summary,
        "audit_report": audit_report,
    }


def _build_factory_mapping(project: Path, source_jsons: list[Path]) -> list[dict[str, str]]:
    """生成厂家名称对照表：报价表厂家名 vs JSON中的厂家名"""
    from openpyxl import load_workbook
    from ocr_price.inventory_color import InventoryColorApplier

    mapping: list[dict[str, str]] = []

    # 1. 读取报价表中的厂家名
    wb = load_workbook(str(project))
    sheet = wb["报价表"]
    applier = InventoryColorApplier(project)

    # 2. 读取JSON文件中的厂家名
    json_factories: set[str] = set()
    for json_path in source_jsons:
        if not json_path.exists():
            continue
        data = json.loads(json_path.read_text(encoding="utf-8"))
        company = data.get("company", "")
        if company:
            json_factories.add(company)

    # 3. 生成对照表
    for name, factory in applier.factory_map.items():
        # 查找匹配的JSON厂家名
        matched_json = None
        for json_name in json_factories:
            if name in json_name or json_name in name:
                matched_json = json_name
                break

        cols = [f"{factory.price_col}列"]
        if factory.extra_cols:
            for prefix, col in factory.extra_cols.items():
                cols.append(f"{col}列({prefix})")

        mapping.append({
            "报价表厂家": name,
            "JSON厂家": matched_json or "未匹配",
            "映射列": ", ".join(cols),
            "状态": "[OK]" if matched_json else "[WARN] 需确认",
        })

    return mapping


def _print_factory_mapping(mapping: list[dict[str, str]]) -> None:
    """打印厂家名称对照表"""
    print("\n" + "=" * 70)
    print("【厂家名称对照表】（请确认映射是否正确）")
    print("=" * 70)
    print(f"{'报价表厂家':<12} {'JSON厂家':<20} {'映射列':<30} {'状态'}")
    print("-" * 70)
    for item in mapping:
        status = item['状态'].replace('\u2713', '[OK]').replace('\u26a0', '[WARN]')
        print(f"{item['报价表厂家']:<12} {item['JSON厂家']:<20} {item['映射列']:<30} {status}")
    print("=" * 70)
    print("请确认以上映射正确后，继续执行...")
    print("=" * 70 + "\n")


def run_single(
    project: Path,
    mode: str,
    list_url: str | None,
    detail_url: str | None,
    account_file: Path,
    username: str | None,
    password: str | None,
    headless: bool,
    image_inputs: list[Path],
    image_jsons: list[Path],
    artifact_dir: Path,
    dry_run: bool = False,
    confirm_write: bool = False,
    refresh_web_artifacts: bool = False,
    refresh_image_artifacts: bool = False,
    manual_login_timeout_seconds: int = 180,
    manifest_path: Path | None = None,
) -> dict[str, Any]:
    artifact_dir.mkdir(parents=True, exist_ok=True)
    project_before_hash = _file_sha256(project) if dry_run and project.exists() else None
    web_location, image_location = _safe_locations(project)
    result: dict[str, Any] = {
        "project": str(project),
        "web_location": web_location,
        "image_location": image_location,
        "mode": mode,
        "started_at": datetime.now().isoformat(timespec="seconds"),
    }

    manifest = _read_manifest(manifest_path) if manifest_path else None
    web_pending = False
    if mode in {"web", "both"}:
        if confirm_write and manifest is not None and not refresh_web_artifacts:
            result["web"] = _web_apply_from_manifest(project=project, manifest=manifest)
        elif confirm_write and not refresh_web_artifacts:
            result["web"] = _web_apply_from_artifacts(
                project=project,
                location=web_location,
                artifact_dir=artifact_dir,
            )
        else:
            result["web"] = _web_flow(
                project=project,
                location=web_location,
                list_url=list_url,
                detail_url=detail_url,
                account_file=account_file,
                username=username,
                password=password,
                headless=headless,
                artifact_dir=artifact_dir,
                apply_if_ready=not dry_run,
                manual_login_timeout_seconds=manual_login_timeout_seconds,
            )
        web_pending = result["web"].get("status") == "pending_confirmation"

    if mode in {"image_doc", "both"}:
        if confirm_write and manifest is not None and not refresh_image_artifacts:
            result["image_doc"] = _image_apply_from_manifest(project=project, manifest=manifest)
        elif confirm_write and not refresh_image_artifacts:
            result["image_doc"] = _image_apply_from_artifacts(
                project=project,
                location=image_location,
                artifact_dir=artifact_dir,
                source_jsons=list(image_jsons),
            )
        else:
            json_sources = list(image_jsons)
            if image_inputs:
                json_sources.extend(_run_ocr_for_inputs(image_inputs, location=_city(image_location), artifact_dir=artifact_dir))

            # 步骤1：生成厂家名称对照表，等待用户确认
            factory_mapping = _build_factory_mapping(project, json_sources)
            if factory_mapping:
                result["factory_mapping"] = factory_mapping
                _print_factory_mapping(factory_mapping)

            image_prepare_only = mode == "both" and web_pending
            result["image_doc"] = _image_flow(
                project=project,
                location=image_location,
                source_jsons=json_sources,
                artifact_dir=artifact_dir,
                apply_if_ready=not image_prepare_only,
                dry_run=dry_run,
            )

    if result.get("web", {}).get("status") == "pending_confirmation" or result.get("image_doc", {}).get("status") == "pending_confirmation":
        result["status"] = "pending_confirmation"
        result["ended_at"] = datetime.now().isoformat(timespec="seconds")
        if dry_run and project_before_hash is not None:
            _assert_dry_run_unchanged(project, project_before_hash)
            result["project_excel_hash_before"] = project_before_hash
            result["project_excel_hash_after"] = _file_sha256(project)
        return result

    result["status"] = "ok"
    result["ended_at"] = datetime.now().isoformat(timespec="seconds")
    if dry_run and project_before_hash is not None:
        _assert_dry_run_unchanged(project, project_before_hash)
        result["project_excel_hash_before"] = project_before_hash
        result["project_excel_hash_after"] = _file_sha256(project)
    return result


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Unified quote update pipeline (single/batch).")
    sub = p.add_subparsers(dest="command", required=True)

    p_single = sub.add_parser("single", help="单文件更新")
    p_single.add_argument("--project", required=True, help="项目Excel路径")
    p_single.add_argument("--mode", choices=("web", "image_doc", "both"), default="both", help="执行模式")
    p_single.add_argument("--list-url", help="网价列表页URL（可选）")
    p_single.add_argument("--detail-url", help="网价详情页URL（可选）")
    p_single.add_argument("--account-file", default="网站账号密码.txt", help="网站账号密码文件")
    p_single.add_argument("--username", help="网站账号")
    p_single.add_argument("--password", help="网站密码")
    p_single.add_argument("--headless", action="store_true", help="网价抓取使用无头浏览器")
    p_single.add_argument("--manual-login-timeout", type=int, default=180, help="有头模式下等待人工登录的秒数")
    p_single.add_argument("--image-inputs", nargs="*", default=[], help="图片/文档原始文件（自动先OCR）")
    p_single.add_argument("--image-jsons", nargs="*", default=[], help="已提取的OCR结果json列表")
    p_single.add_argument("--artifact-dir", default="运行产物", help="产物目录")
    p_single.add_argument("--dry-run", action="store_true", help="预演流程，只生成报告，不修改项目Excel")
    p_single.add_argument("--confirm-write", action="store_true", help="明确允许写入项目Excel")
    p_single.add_argument("--refresh-web-artifacts", action="store_true", help="confirm-write时强制重跑网价抓取与对照生成")
    p_single.add_argument("--refresh-image-artifacts", action="store_true", help="confirm-write时强制重跑OCR与图片文档对照生成")
    p_single.add_argument("--manifest", help="dry-run生成的manifest路径；confirm-write默认必须复用它")
    p_single.add_argument("--report-out", help="单文件总结报告路径")

    p_batch = sub.add_parser("batch", help="批量更新")
    p_batch.add_argument("--project-dir", required=True, help="项目Excel目录")
    p_batch.add_argument("--glob", default="*.xlsx", help="项目Excel匹配模式")
    p_batch.add_argument("--mode", choices=("web", "image_doc", "both"), default="both", help="执行模式")
    p_batch.add_argument("--list-url", help="网价列表页URL（可选）")
    p_batch.add_argument("--detail-url", help="网价详情页URL（可选）")
    p_batch.add_argument("--account-file", default="网站账号密码.txt", help="网站账号密码文件")
    p_batch.add_argument("--username", help="网站账号")
    p_batch.add_argument("--password", help="网站密码")
    p_batch.add_argument("--headless", action="store_true", help="网价抓取使用无头浏览器")
    p_batch.add_argument("--manual-login-timeout", type=int, default=180, help="有头模式下等待人工登录的秒数")
    p_batch.add_argument("--image-source-map", help="批量图片源映射json：{项目文件名:[json或原始文件...]}")
    p_batch.add_argument("--artifact-dir", default="运行产物", help="产物目录")
    p_batch.add_argument("--dry-run", action="store_true", help="预演流程，只生成报告，不修改项目Excel")
    p_batch.add_argument("--confirm-write", action="store_true", help="明确允许写入项目Excel")
    p_batch.add_argument("--refresh-web-artifacts", action="store_true", help="confirm-write时强制重跑网价抓取与对照生成")
    p_batch.add_argument("--refresh-image-artifacts", action="store_true", help="confirm-write时强制重跑OCR与图片文档对照生成")
    p_batch.add_argument("--manifest", help="dry-run生成的manifest路径；单文件建议使用，批量后续拆分支持")
    p_batch.add_argument("--report-out", help="批量总结报告路径")
    return p


def _split_image_sources(items: list[str]) -> tuple[list[Path], list[Path]]:
    inputs: list[Path] = []
    jsons: list[Path] = []
    for x in items:
        p = Path(x)
        if p.suffix.lower() == ".json":
            jsons.append(p)
        else:
            inputs.append(p)
    return inputs, jsons


def _default_offline_sources(offline_dir: Path = DEFAULT_OFFLINE_DIR) -> list[str]:
    if not offline_dir.exists() or not offline_dir.is_dir():
        return []
    return [
        str(p)
        for p in sorted(offline_dir.iterdir())
        if p.is_file() and p.suffix.lower() in OFFLINE_SOURCE_SUFFIXES
    ]


def main() -> int:
    args = _build_parser().parse_args()
    if args.dry_run and args.confirm_write:
        raise SystemExit("--dry-run 与 --confirm-write 不能同时使用。")
    if not args.dry_run and not args.confirm_write:
        raise SystemExit("为防止误操作，写入项目Excel必须显式传入 --confirm-write；预演请使用 --dry-run。")
    artifact_base_dir = Path(args.artifact_dir)
    artifact_base_dir.mkdir(parents=True, exist_ok=True)

    if args.command == "single":
        project_path = Path(args.project)
        project_artifact_dir = _project_artifact_dir(artifact_base_dir, project_path)
        lock_path = artifact_base_dir / ".quote_update.lock"
        image_inputs = [Path(x) for x in args.image_inputs]
        image_jsons = [Path(x) for x in args.image_jsons]
        auto_collect_sources = (
            args.mode in {"image_doc", "both"}
            and (args.dry_run or args.refresh_image_artifacts)
            and not image_inputs
            and not image_jsons
        )
        if auto_collect_sources:
            auto_sources = _default_offline_sources()
            auto_inputs, auto_jsons = _split_image_sources(auto_sources)
            image_inputs = auto_inputs
            image_jsons = auto_jsons
        with _pipeline_lock(lock_path, project_path):
            result = run_single(
                project=project_path,
                mode=args.mode,
                list_url=args.list_url,
                detail_url=args.detail_url,
                account_file=Path(args.account_file),
                username=args.username,
                password=args.password,
                headless=args.headless,
                image_inputs=image_inputs,
                image_jsons=image_jsons,
                artifact_dir=project_artifact_dir,
                dry_run=args.dry_run,
                confirm_write=args.confirm_write,
                refresh_web_artifacts=args.refresh_web_artifacts,
                refresh_image_artifacts=args.refresh_image_artifacts,
                manual_login_timeout_seconds=args.manual_login_timeout,
                manifest_path=Path(args.manifest) if args.manifest else None,
            )
            if args.dry_run:
                manifest_out = _manifest_path(project_artifact_dir)
                _write_manifest(manifest_out, result, project_artifact_dir)
                print(f"Manifest: {manifest_out}")
            report_out = (
                Path(args.report_out)
                if args.report_out
                else project_artifact_dir / f"完整更新报告_single_{project_path.stem}_{_ts()}.json"
            )
            _json_write(report_out, result)
            markdown_report = render_single_report_markdown(result, json_report_path=str(report_out))
            markdown_out = report_out.with_suffix(".md")
            markdown_out.write_text(markdown_report, encoding="utf-8")
            print(f"Status: {result['status']}")
            print(f"Report: {report_out}")
            print(f"MarkdownReport: {markdown_out}")
            _print_result_details(result)
        return 0

    source_map = _json_read(Path(args.image_source_map)) if args.image_source_map else {}
    if source_map and not isinstance(source_map, dict):
        raise SystemExit("--image-source-map 必须是对象：{项目文件名:[source1,source2]}")
    default_offline_sources = _default_offline_sources()

    projects = sorted(Path(args.project_dir).glob(args.glob))
    summary: dict[str, Any] = {
        "mode": args.mode,
        "project_dir": str(Path(args.project_dir)),
        "glob": args.glob,
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "results": [],
    }
    ok = 0
    pending = 0
    failed = 0

    for project in projects:
        if project.name.startswith("~$"):
            continue
        try:
            if isinstance(source_map, dict) and source_map:
                mapped = source_map.get(project.name, [])
            else:
                should_collect = args.mode in {"image_doc", "both"} and (args.dry_run or args.refresh_image_artifacts)
                mapped = default_offline_sources if should_collect else []
            img_inputs, img_jsons = _split_image_sources(mapped)
            project_artifact_dir = _project_artifact_dir(artifact_base_dir, project)
            result = run_single(
                project=project,
                mode=args.mode,
                list_url=args.list_url,
                detail_url=args.detail_url,
                account_file=Path(args.account_file),
                username=args.username,
                password=args.password,
                headless=args.headless,
                image_inputs=img_inputs,
                image_jsons=img_jsons,
                artifact_dir=project_artifact_dir,
                dry_run=args.dry_run,
                confirm_write=args.confirm_write,
                refresh_web_artifacts=args.refresh_web_artifacts,
                refresh_image_artifacts=args.refresh_image_artifacts,
                manual_login_timeout_seconds=args.manual_login_timeout,
            )
            summary["results"].append(result)
            _print_result_details(result)
            if result.get("status") == "ok":
                ok += 1
            elif result.get("status") == "pending_confirmation":
                pending += 1
            else:
                failed += 1
        except Exception as exc:  # noqa: BLE001
            failed += 1
            summary["results"].append(
                {
                    "project": str(project),
                    "status": "failed",
                    "error": f"{type(exc).__name__}: {exc}",
                }
            )

    summary["ok_count"] = ok
    summary["pending_confirmation_count"] = pending
    summary["failed_count"] = failed
    summary["ended_at"] = datetime.now().isoformat(timespec="seconds")
    report_out = Path(args.report_out) if args.report_out else artifact_base_dir / f"完整更新报告_batch_{_ts()}.json"
    _json_write(report_out, summary)
    print(f"Projects: {len(summary['results'])}")
    print(f"OK: {ok}, Pending: {pending}, Failed: {failed}")
    print(f"Report: {report_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
