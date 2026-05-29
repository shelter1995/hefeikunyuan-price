
from __future__ import annotations

import argparse
import base64
import csv
import json
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

from .xlsx_utils import load_workbook_safe

SKIP_SHEETS = {"报价表"}
MAPPING_HEADERS = ["项目文件Sheet", "最新清单厂家Sheet", "状态", "说明"]
CONFIRMED_WRITE_STATUS = "已确认匹配"
CONFIRMED_SKIP_STATUS = "已确认不更新"

LOCATION_KEYWORDS = [
    "安徽蚌埠", "安徽合肥", "江苏南通",
    "安徽阜阳", "安徽宿州", "安徽淮南", "安徽芜湖", "安徽马鞍山", "安徽铜陵",
    "江苏徐州", "江苏宿迁", "江苏南京", "江苏苏州", "江苏无锡", "江苏常州",
    "浙江杭州", "浙江宁波", "浙江温州",
    "山东济南", "山东青岛", "山东临沂", "山东日照", "山东莱芜",
    "河南郑州", "河南洛阳",
    "湖北武汉", "湖北宜昌",
    "上海",
    "蚌埠", "合肥", "南通", "徐州", "宿迁", "阜阳", "芜湖", "南京",
]
LOCATION_LIST_URLS = {
    "安徽蚌埠": "https://jiancai.mysteel.com/market/pa228aa010101a0a01010715aaaa1.html",
    "安徽合肥": "https://jiancai.mysteel.com/market/pa228aa010101a0a01010701aaaa1.html",
    "江苏南通": "https://jiancai.mysteel.com/market/pa228aa010101a0a01010308aaaa1.html",
    "蚌埠": "https://jiancai.mysteel.com/market/pa228aa010101a0a01010715aaaa1.html",
    "合肥": "https://jiancai.mysteel.com/market/pa228aa010101a0a01010701aaaa1.html",
    "南通": "https://jiancai.mysteel.com/market/pa228aa010101a0a01010308aaaa1.html",
}

ROW_HEADERS = ["品名", "规格(mm)", "材质", "钢厂/产地", "价格(元/吨)", "涨跌", "备注", "钢号"]
STOPWORDS = ("安徽", "江苏", "河南", "集团", "钢铁", "钢厂", "贸易", "有限公司", "有限责任公司", "公司")
INVALID_SHEET_CHARS = {"\\": "＼", "/": "／", "*": "＊", "[": "【", "]": "】", ":": "：", "?": "？"}


class WebPriceError(RuntimeError):
    pass


def _n(x: str) -> str:
    return re.sub(r"\s+", "", x or "")


def _clean_place_text(value: str) -> str:
    text = str(value or "").replace("\u3000", " ").replace("\xa0", " ").strip()
    if not text:
        return ""
    # 某些行会把说明前缀和厂家名放在同一个单元格的多行文本里，取最后一行做厂家名
    parts = [p.strip() for p in re.split(r"[\r\n]+", text) if p.strip()]
    text = parts[-1] if parts else text
    text = re.sub(r"\s*/\s*", "/", text)
    return _n(text)


def _norm_company(name: str) -> str:
    value = _n(name)
    for token in STOPWORDS:
        value = value.replace(token, "")
    return value


def _parse_location(filename: str) -> tuple[str, str]:
    """
    Parse dual-location from filename like '地名1-地名2-XXX'.
    Returns (web_location, image_doc_location).
    For backward compatibility, if only one location found,
    both web and image_doc use the same location.
    """
    stem = Path(filename).stem
    found: list[str] = []
    keywords = sorted(set(LOCATION_KEYWORDS), key=len, reverse=True)

    # Scan from left to right and choose the longest keyword at each position.
    # This preserves visual order in filename and avoids overlap matches
    # (e.g. "安徽合肥" should not also produce an extra "合肥" at same span).
    i = 0
    while i < len(stem):
        matched = None
        for keyword in keywords:
            if stem.startswith(keyword, i):
                matched = keyword
                break
        if matched:
            found.append(matched)
            i += len(matched)
        else:
            i += 1

    if not found:
        raise WebPriceError(f"无法从文件名识别地点: {filename}")
    if len(found) >= 2:
        return found[0], found[1]
    return found[0], found[0]


def _city(location: str) -> str:
    for p in ("安徽", "江苏"):
        if location.startswith(p):
            return location[len(p) :]
    return location


def _parse_date(text: str) -> str | None:
    m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text or "")
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text or "")
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.search(r"(\d{1,2})月(\d{1,2})日", text or "")
    if m:
        y = datetime.now().year
        return f"{y:04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None


def _sort_key_for_latest(text: str, href: str) -> tuple[int, str]:
    date_text = _parse_date(text)
    hh = 0
    mm = 0
    tm = re.search(r"\((\d{1,2}):(\d{2})\)", text or "")
    if tm:
        hh = int(tm.group(1))
        mm = int(tm.group(2))
    if date_text:
        return (int(date_text.replace("-", "") + f"{hh:02d}{mm:02d}"), text)

    # href样式通常包含 yymmddhh 或 yymmdd
    hm = re.search(r"/m/(\d{6})(\d{2})?/", href or "")
    if hm:
        yymmdd = hm.group(1)
        hour = hm.group(2) or "00"
        return (int(f"20{yymmdd}{hour}00"), text)
    return (0, text)


def _price(v: Any) -> int | None:
    if v is None:
        return None
    m = re.search(r"(?<!\d)(\d{3,5})(?!\d)", str(v).replace(",", ""))
    return int(m.group(1)) if m else None


def _score_company(left: str, right: str) -> int:
    l = _norm_company(left)
    r = _norm_company(right)
    if not l or not r:
        return 0
    if l == r:
        return 100
    if l.startswith(r) or r.startswith(l):
        return 80
    if l in r or r in l:
        return 60
    if len(l) >= 2 and len(r) >= 2 and l[:2] == r[:2]:
        return 30
    return 0


def _score_10(spec: str) -> int:
    s = _n(spec).replace("φ", "Φ")
    if re.fullmatch(r"(?:Φ)?8[-~～]10", s):
        return 400
    if re.fullmatch(r"(?:Φ)?10(?:mm)?", s):
        return 300
    for m in re.finditer(r"(\d{1,2})\s*[-~～]\s*(\d{1,2})", s):
        if int(m.group(1)) <= 10 <= int(m.group(2)):
            return 200
    return 100 if "10" in s else 0


def _score_18(spec: str) -> int:
    s = _n(spec).replace("φ", "Φ")
    if re.fullmatch(r"(?:Φ)?18(?:mm)?", s):
        return 200
    return 100 if "18" in s else 0


def _safe_sheet(name: str, used: set[str]) -> str:
    n = (name or "未知钢厂").strip()
    for a, b in INVALID_SHEET_CHARS.items():
        n = n.replace(a, b)
    n = n[:31] or "未知钢厂"
    if n not in used:
        used.add(n)
        return n
    i = 2
    while True:
        s = f"_{i}"
        c = f"{n[:31-len(s)]}{s}"
        if c not in used:
            used.add(c)
            return c
        i += 1


def _parse_credentials(account_file: Path, username: str | None, password: str | None) -> tuple[str, str]:
    if username and password:
        return username, password
    text = account_file.read_text(encoding="utf-8", errors="ignore") if account_file.exists() else ""
    u = username
    p = password
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if ("账号" in s or "用户名" in s or "user" in s.lower()) and ("：" in s or ":" in s):
            u = u or s.split("：")[-1].split(":")[-1].strip()
        if ("密码" in s or "pass" in s.lower()) and ("：" in s or ":" in s):
            p = p or s.split("：")[-1].split(":")[-1].strip()
    if not u or not p:
        vals = [x.strip() for x in text.splitlines() if x.strip()]
        if len(vals) >= 2:
            u = u or vals[0].split()[-1]
            p = p or vals[1].split()[-1]
    if not u or not p:
        raise WebPriceError("无法解析账号密码，请检查网站账号密码.txt或传入--username/--password")
    return u, p


def _write_mapping_json_csv(mapping_rows: list[dict[str, str]], json_path: Path, csv_path: Path) -> None:
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(mapping_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MAPPING_HEADERS)
        writer.writeheader()
        for row in mapping_rows:
            writer.writerow({k: row.get(k, "") for k in MAPPING_HEADERS})


def _load_mapping_rows(path: Path) -> list[dict[str, str]]:
    """加载映射行，支持待确认和已确认文件自动识别"""
    if not path.exists():
        # 尝试查找已确认文件
        confirmed_path = Path(str(path).replace("待确认", "已确认"))
        if confirmed_path.exists():
            path = confirmed_path
        else:
            return []
    rows = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(rows, list):
        raise WebPriceError(f"映射文件格式错误（应为数组）: {path}")
    out: list[dict[str, str]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        out.append(
            {
                "项目文件Sheet": str(row.get("项目文件Sheet") or "").strip(),
                "最新清单厂家Sheet": str(row.get("最新清单厂家Sheet") or "").strip(),
                "状态": str(row.get("状态") or "").strip(),
                "说明": str(row.get("说明") or "").strip(),
            }
        )
    return out


def _mapping_status_priority(status: str) -> int:
    s = str(status or "").strip()
    if s == CONFIRMED_WRITE_STATUS:
        return 50
    if s == CONFIRMED_SKIP_STATUS:
        return 40
    if s.startswith("待确认"):
        return 30
    if s.startswith("未匹配"):
        return 20
    if s.startswith("跳过"):
        return 10
    return 0


def _mapping_row_key(row: dict[str, str]) -> tuple[str, str] | None:
    sheet = str(row.get("项目文件Sheet") or "").strip()
    source = str(row.get("最新清单厂家Sheet") or "").strip()
    if sheet:
        return ("sheet", sheet)
    if source:
        norm = _norm_company(source) or source
        return ("source", norm)
    return None


def _mapping_row_score(row: dict[str, str], idx: int) -> tuple[int, int, int, int]:
    source = str(row.get("最新清单厂家Sheet") or "").strip()
    note = str(row.get("说明") or "").strip()
    return (_mapping_status_priority(str(row.get("状态") or "")), 1 if source else 0, 1 if note else 0, idx)


def _dedupe_mapping_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    if not rows:
        return rows
    best_idx_by_key: dict[tuple[str, str], int] = {}
    best_score_by_key: dict[tuple[str, str], tuple[int, int, int, int]] = {}
    keep_nonkey: set[int] = set()
    for idx, row in enumerate(rows):
        if not isinstance(row, dict):
            keep_nonkey.add(idx)
            continue
        key = _mapping_row_key(row)
        if key is None:
            keep_nonkey.add(idx)
            continue
        score = _mapping_row_score(row, idx)
        prev_score = best_score_by_key.get(key)
        if prev_score is None or score > prev_score:
            best_score_by_key[key] = score
            best_idx_by_key[key] = idx
    keep_idx = keep_nonkey | set(best_idx_by_key.values())
    out: list[dict[str, str]] = []
    for idx, row in enumerate(rows):
        if idx in keep_idx and isinstance(row, dict):
            out.append(row)
    return out


def _project_sheets(project_excel: Path) -> list[str]:
    wb = load_workbook_safe(project_excel, data_only=True)
    return [x for x in wb.sheetnames if x not in SKIP_SHEETS]


def _export_raw(rows: list[dict[str, Any]], output_excel: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    by_mill: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        by_mill.setdefault(_n(str(r.get("钢厂/产地") or "")), []).append(r)
    used: set[str] = set()
    for mill in sorted(by_mill):
        if not mill:
            continue
        ws = wb.create_sheet(_safe_sheet(mill, used))
        ws.append(ROW_HEADERS)
        for r in by_mill[mill]:
            ws.append([r.get(k) for k in ROW_HEADERS])
    output_excel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_excel)


def _load_raw(raw_excel: Path) -> list[dict[str, Any]]:
    wb = load_workbook_safe(raw_excel, data_only=True)
    out: list[dict[str, Any]] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        heads = [str(x.value or "").strip() for x in ws[1]]
        idx = {h: i for i, h in enumerate(heads)}
        if not all(k in idx for k in ["品名", "规格(mm)", "材质", "钢厂/产地", "价格(元/吨)"]):
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            mill = _clean_place_text(str(row[idx["钢厂/产地"]] or ""))
            if not mill:
                continue
            out.append(
                {
                    "品名": str(row[idx["品名"]] or "").strip(),
                    "规格(mm)": str(row[idx["规格(mm)"]] or "").strip(),
                    "材质": str(row[idx["材质"]] or "").strip(),
                    "钢厂/产地": mill,
                    "价格(元/吨)": _price(row[idx["价格(元/吨)"]]),
                    "涨跌": str(row[idx.get("涨跌", -1)] or "").strip() if "涨跌" in idx else "",
                    "备注": str(row[idx.get("备注", -1)] or "").strip() if "备注" in idx else "",
                    "钢号": str(row[idx.get("钢号", -1)] or "").strip() if "钢号" in idx else "",
                }
            )
    return out


def _source_mills(rows: list[dict[str, Any]]) -> list[str]:
    return sorted(
        {
            _clean_place_text(str(r.get("钢厂/产地") or ""))
            for r in rows
            if _clean_place_text(str(r.get("钢厂/产地") or ""))
        }
    )

def _match_source_name(name: str, source_mills: list[str]) -> str | None:
    if not name:
        return None
    n = _norm_company(name)
    if not n:
        return None
    for src in source_mills:
        if _norm_company(src) == n:
            return src
    scored = []
    for src in source_mills:
        score = _score_company(name, src)
        if score > 0:
            scored.append((score, src))
    if not scored:
        return None
    scored.sort(key=lambda x: x[0], reverse=True)
    top = [s for sc, s in scored if sc == scored[0][0]]
    return top[0] if len(top) == 1 else None


def _suggest_source_for_sheet(sheet: str, source_mills: list[str]) -> tuple[str, str, str]:
    scored: list[tuple[int, str]] = []
    for src in source_mills:
        score = _score_company(sheet, src)
        if score > 0:
            scored.append((score, src))
    if not scored:
        return "", "未匹配(不更新)", "最新清单厂家中无对应项"
    scored.sort(key=lambda x: x[0], reverse=True)
    top_score = scored[0][0]
    top = [src for score, src in scored if score == top_score]
    if len(top) > 1:
        return "", "待确认(冲突)", "候选: " + " / ".join(sorted(set(top)))
    return top[0], "待确认匹配", f"自动建议（score={top_score}），需人工确认"


def _mapping_confirmed_sources(rows: list[dict[str, str]]) -> set[str]:
    out: set[str] = set()
    for row in rows:
        status = str(row.get("状态") or "").strip()
        source = str(row.get("最新清单厂家Sheet") or "").strip()
        if status in {CONFIRMED_WRITE_STATUS, CONFIRMED_SKIP_STATUS} and source:
            out.add(_norm_company(source))
    return out


def _mapping_mapped_sources(rows: list[dict[str, str]]) -> set[str]:
    out: set[str] = set()
    for row in rows:
        source = str(row.get("最新清单厂家Sheet") or "").strip()
        if source:
            out.add(_norm_company(source))
    return out


def _mapping_project_set(rows: list[dict[str, str]]) -> set[str]:
    return {
        str(r.get("项目文件Sheet") or "").strip()
        for r in rows
        if str(r.get("项目文件Sheet") or "").strip() and str(r.get("项目文件Sheet") or "").strip() not in SKIP_SHEETS
    }


def _mapping_source_set(rows: list[dict[str, str]]) -> set[str]:
    return {
        _norm_company(str(r.get("最新清单厂家Sheet") or "").strip())
        for r in rows
        if str(r.get("最新清单厂家Sheet") or "").strip()
    }


def _pick_coil_10(rows: list[dict[str, Any]]) -> int | None:
    best_score = -1
    best_price: int | None = None
    for row in rows:
        name = _n(str(row.get("品名") or ""))
        mat = _n(str(row.get("材质") or ""))
        if "盘螺" not in name or mat != "HRB400E":
            continue
        price = _price(row.get("价格(元/吨)"))
        if price is None:
            continue
        score = _score_10(str(row.get("规格(mm)") or ""))
        if score > best_score:
            best_score = score
            best_price = price
    return best_price if best_score > 0 else None


def _score_14(spec: str) -> int:
    s = _n(spec).replace("φ", "Φ")
    if re.fullmatch(r"(?:Φ)?14(?:mm)?", s):
        return 200
    return 100 if "14" in s else 0


def _pick_rebar_for_mill(rows: list[dict[str, Any]], mill_name: str) -> int | None:
    """
    Pick rebar price: 金虹 uses Φ14, others use Φ18.
    """
    norm_mill = _norm_company(mill_name)
    use_14 = "金虹" in norm_mill or "金虹" in mill_name

    best_score = -1
    best_price: int | None = None
    for row in rows:
        name = _n(str(row.get("品名") or ""))
        mat = _n(str(row.get("材质") or ""))
        if "螺纹钢" not in name or mat != "HRB400E":
            continue
        price = _price(row.get("价格(元/吨)"))
        if price is None:
            continue
        spec = str(row.get("规格(mm)") or "")
        score = _score_14(spec) if use_14 else _score_18(spec)
        if score > best_score:
            best_score = score
            best_price = price
    return best_price if best_score > 0 else None


def _click_any(page: Any, selectors: list[str], timeout_ms: int = 5000) -> bool:
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count() > 0:
                loc.click(timeout=timeout_ms)
                return True
        except Exception:
            continue
    return False


def _visible_any(page: Any, selectors: list[str]) -> bool:
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count() > 0 and loc.is_visible():
                return True
        except Exception:
            continue
    return False


def _strip_jsonp(payload: str) -> str:
    text = (payload or "").strip()
    m = re.match(r"^[^(]*\((.*)\)\s*;?\s*$", text, re.S)
    return m.group(1).strip() if m else text


def _decode_userinfo_blob(value: str) -> dict[str, Any]:
    s = str(value or "").strip()
    if not s:
        return {}
    try:
        padded = s + "=" * ((4 - len(s) % 4) % 4)
        raw = base64.b64decode(padded)
        obj = json.loads(raw.decode("utf-8", errors="ignore"))
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}


def _fetch_userinfo(page: Any) -> dict[str, Any]:
    try:
        resp = page.request.get(
            f"https://passport.mysteel.com/api/userinfo.htm?callback=callback&_={int(datetime.now().timestamp() * 1000)}",
            timeout=15000,
        )
        text = _strip_jsonp(resp.text())
        payload = json.loads(text) if text else {}
        if not isinstance(payload, dict):
            return {}
        decoded = _decode_userinfo_blob(str(payload.get("info") or ""))
        if decoded:
            payload.update(decoded)
        return payload
    except Exception:
        return {}


def _userinfo_logged_in(payload: dict[str, Any]) -> bool:
    if not isinstance(payload, dict):
        return False
    for key in ("loginMobile", "username", "memberName", "phone"):
        value = payload.get(key)
        if isinstance(value, str) and value.strip():
            return True
    for key in ("userId", "memberId"):
        value = payload.get(key)
        if isinstance(value, (int, float)) and int(value) > 0:
            return True
    return False


def _detect_logged_in(page: Any, retry: bool = False) -> tuple[bool, str]:
    if _visible_any(
        page,
        [
            "a:has-text('退出')",
            ".topbar-nav-logout:visible",
            ".topbar-user-name:visible",
            ".topbar-nav-news:visible",
            ".topbar-nav-msg:visible",
        ],
    ):
        return True, "顶部导航已显示登录态标记"
    info = _fetch_userinfo(page)
    if _userinfo_logged_in(info):
        return True, "userinfo接口返回已登录"
    if not retry:
        page.wait_for_timeout(2000)
        return _detect_logged_in(page, retry=True)
    return False, "未检测到登录态"


def _wait_for_manual_login(
    page: Any,
    timeout_seconds: int,
    poll_interval_seconds: int,
) -> tuple[bool, str]:
    timeout_seconds = max(1, timeout_seconds)
    poll_interval_seconds = max(1, poll_interval_seconds)
    attempts = max(1, timeout_seconds // poll_interval_seconds)
    print(
        "[web_price] 自动登录未成功。请在打开的浏览器中手动完成登录，"
        f"脚本将等待最多 {timeout_seconds} 秒。"
    )
    for _ in range(attempts):
        page.wait_for_timeout(poll_interval_seconds * 1000)
        ok, proof = _detect_logged_in(page)
        if ok:
            return True, proof
    return False, "等待人工登录超时，仍未检测到登录态"


def _login(
    page: Any,
    user: str,
    pwd: str,
    allow_manual_login: bool = False,
    manual_login_timeout_seconds: int = 180,
    manual_login_poll_interval_seconds: int = 3,
) -> str:
    page.goto("https://www.mysteel.com/", timeout=60000, wait_until="domcontentloaded")
    page.wait_for_timeout(3000)
    already_logged, reason = _detect_logged_in(page)
    if already_logged:
        return f"已登录，跳过登录（{reason}）"
    need_login = page.locator(".topbar-nav-login:visible").count() > 0
    if need_login:
        opened = False
        for _ in range(3):
            if not opened:
                opened = _click_any(
                    page,
                    [".topbar-nav-login:visible", ".topbar-nav-login", "text=登录", "a:has-text('登录')"],
                    timeout_ms=15000,
                )
            _click_any(
                page,
                [".form-tab-account:visible", ".form-tab-account", "text=账号登录", "button:has-text('账号登录')"],
                timeout_ms=8000,
            )
            page.wait_for_timeout(800)
            if page.locator('input[placeholder="请输入用户名"]:visible').count() > 0:
                break
        if not opened:
            if not allow_manual_login:
                raise WebPriceError("登录入口未找到（topbar-nav-login/登录按钮）")
            print("[web_price] 自动点击登录入口失败，改为等待人工登录。")

        user_input = page.locator('input[placeholder="请输入用户名"]:visible').first
        pwd_input = page.locator('input[placeholder="请输入密码"]:visible').first
        if (user_input.count() == 0 or pwd_input.count() == 0) and not allow_manual_login:
            raise WebPriceError("账号登录输入框不可见，请检查登录弹窗状态")
        if user_input.count() > 0 and pwd_input.count() > 0:
            user_input.fill(user)
            pwd_input.fill(pwd)
        elif allow_manual_login:
            print("[web_price] 自动登录输入框不可见，改为等待人工登录。")
        if user_input.count() > 0 and pwd_input.count() > 0 and not _click_any(
            page,
            [".form-button-login:visible", ".form-button-login", "button:has-text('登录')"],
            timeout_ms=10000,
        ):
            if not allow_manual_login:
                raise WebPriceError("登录按钮未找到")
            print("[web_price] 自动点击登录按钮失败，改为等待人工登录。")
    page.wait_for_timeout(4000)
    ok, proof = _detect_logged_in(page)
    if not ok:
        page.wait_for_timeout(2000)
        ok, proof = _detect_logged_in(page)
    if not ok and allow_manual_login:
        ok, proof = _wait_for_manual_login(
            page,
            timeout_seconds=manual_login_timeout_seconds,
            poll_interval_seconds=manual_login_poll_interval_seconds,
        )
        if ok:
            return f"人工登录成功（{proof}）"
    if not ok:
        raise WebPriceError(f"登录失败：{proof}。如遇验证码/滑块，请去掉 --headless 使用有头浏览器手动登录。")
    return f"登录成功（{proof}）"


def _latest_url_from_list(page: Any, list_url: str, city: str) -> str:
    page.goto(list_url, timeout=60000, wait_until="domcontentloaded")
    page.wait_for_timeout(3000)
    title = page.title()
    body = page.locator("body").inner_text()
    if "安全验证" in title or "操作异常" in body or "EventID" in body:
        raise WebPriceError("列表页触发安全验证，无法自动定位最新报价。请使用 --detail-url 显式传入。")

    items = page.eval_on_selector_all(
        "a",
        """
        (els) => els.map(e => ({text:(e.innerText||'').trim(), href:e.href||''}))
            .filter(x => x.href.includes('/m/') && x.href.endsWith('.html'))
        """,
    )
    cands: list[tuple[str, str]] = []
    for it in items:
        t = (it.get("text") or "").strip()
        h = (it.get("href") or "").strip()
        if not t or not h:
            continue
        if city in t and "建筑钢材" in t and ("行情" in t or "报价" in t):
            cands.append((h, t))
    if not cands:
        for it in items:
            t = (it.get("text") or "").strip()
            h = (it.get("href") or "").strip()
            if city in t and "建筑钢材" in t:
                cands.append((h, t))
    if not cands:
        raise WebPriceError("未在列表页定位到目标城市报价链接，请使用 --detail-url 显式传入。")
    cands.sort(key=lambda x: _sort_key_for_latest(x[1], x[0]), reverse=True)
    return cands[0][0]


def _extract_detail_rows(page: Any) -> list[dict[str, Any]]:
    try:
        page.wait_for_selector("#marketTable", timeout=45000)
    except PlaywrightTimeoutError as exc:
        raise WebPriceError("详情页未加载出 marketTable") from exc
    page.wait_for_timeout(2000)
    if page.locator(".icon-key").count() > 0:
        raise WebPriceError("详情页价格仍为加密态（icon-key存在），请检查登录状态")

    rows = page.eval_on_selector_all(
        "#marketTable tbody tr",
        """
        (els) => els.map(tr => Array.from(tr.querySelectorAll('td')).map(td => ({
            type: td.getAttribute('data-type') || '',
            text: (td.innerText || '').trim()
        })))
        """,
    )

    out: list[dict[str, Any]] = []
    for row in rows:
        m: dict[str, str] = {}
        plain: list[str] = []
        for c in row:
            if not isinstance(c, dict):
                continue
            t = str(c.get("type") or "").strip()
            v = str(c.get("text") or "").strip()
            if t:
                m[t] = v
            plain.append(v)

        rec = {
            "品名": m.get("breed", plain[0] if len(plain) > 0 else ""),
            "规格(mm)": m.get("spec", plain[1] if len(plain) > 1 else ""),
            "材质": m.get("material", plain[2] if len(plain) > 2 else ""),
            "钢厂/产地": _clean_place_text(m.get("place", plain[3] if len(plain) > 3 else "")),
            "价格(元/吨)": _price(m.get("price", plain[4] if len(plain) > 4 else "")),
            "涨跌": m.get("raise", plain[5] if len(plain) > 5 else ""),
            "备注": m.get("note", plain[6] if len(plain) > 6 else ""),
            "钢号": m.get("GANGHAO", plain[7] if len(plain) > 7 else ""),
        }
        if rec["钢厂/产地"]:
            out.append(rec)

    if not out:
        raise WebPriceError("详情页未解析到有效报价行")
    return out


def _get_user_data_dir() -> Path:
    """获取 Chrome 用户数据目录，用于保持登录状态。"""
    return Path(__file__).parent.parent / ".chrome_user_data"


def _launch_browser_with_state(p, headless: bool):
    """启动浏览器，使用持久化用户数据目录保持登录状态。

    返回 (browser, context, page, connection_type)
    """
    user_data_dir = _get_user_data_dir()
    user_data_dir.mkdir(parents=True, exist_ok=True)

    # 尝试通过 CDP 连接已运行的 Chrome（Chrome MCP 等）
    try:
        browser = p.chromium.connect_over_cdp("http://localhost:9222")
        contexts = browser.contexts
        context = contexts[0] if contexts else browser.new_context()
        page = context.pages[0] if context.pages else context.new_page()
        print(f"[web_price] 通过 CDP 连接到已运行的 Chrome 浏览器")
        return browser, context, page, "cdp"
    except Exception:
        pass

    # 使用持久化用户数据目录启动浏览器
    print(f"[web_price] 使用持久化用户数据目录启动浏览器: {user_data_dir}")
    browser = p.chromium.launch_persistent_context(
        user_data_dir=str(user_data_dir),
        headless=headless,
        locale="zh-CN",
        args=["--disable-blink-features=AutomationControlled"],
    )
    page = browser.pages[0] if browser.pages else browser.new_page()
    return browser, browser, page, "persistent"


def fetch_web_prices(
    project_excel: Path,
    location: str | None,
    list_url: str | None,
    detail_url: str | None,
    account_file: Path,
    username: str | None,
    password: str | None,
    output_excel: Path,
    report_out: Path,
    headless: bool,
    manual_login_timeout_seconds: int = 180,
) -> dict[str, Any]:
    loc = location or _parse_location(project_excel.name)[0]
    city = _city(loc)
    user, pwd = _parse_credentials(account_file, username, password)
    list_page = list_url or LOCATION_LIST_URLS.get(loc)
    if not detail_url and not list_page:
        raise WebPriceError(f"未配置地点列表页URL: {loc}，请传 --list-url 或 --detail-url")

    with sync_playwright() as p:
        browser, context, page, connection_type = _launch_browser_with_state(p, headless)

        try:
            proof = _login(
                page,
                user,
                pwd,
                allow_manual_login=not headless,
                manual_login_timeout_seconds=manual_login_timeout_seconds,
            )
        except WebPriceError:
            if connection_type == "cdp":
                raise
            print(
                "[web_price] 无头模式登录失败，将关闭并重新启动有头浏览器，"
                "请在弹出的浏览器窗口中手动完成登录。"
            )
            browser.close()
            browser, context, page, connection_type = _launch_browser_with_state(p, headless=False)
            proof = _login(
                page,
                user,
                pwd,
                allow_manual_login=True,
                manual_login_timeout_seconds=manual_login_timeout_seconds,
            )

        url = detail_url or _latest_url_from_list(page, str(list_page), city)
        page.goto(url, timeout=60000, wait_until="networkidle")
        rows = _extract_detail_rows(page)
        h1 = page.locator("h1").first.inner_text().strip() if page.locator("h1").count() else ""
        title = page.title()
        qdate = _parse_date(h1) or _parse_date(title) or datetime.now().strftime("%Y-%m-%d")

        # 只有非持久化上下文才需要关闭（persistent context 关闭会保存状态）
        if connection_type == "cdp":
            # CDP 连接不关闭浏览器
            pass
        else:
            # persistent context 关闭时会保存 cookie 和 localStorage
            browser.close()

    _export_raw(rows, output_excel)
    report = {
        "phase": "web_fetch",
        "project_excel": project_excel.name,
        "location": loc,
        "city": city,
        "list_url": list_page,
        "latest_url": url,
        "quote_date": qdate,
        "title": title,
        "login_proof": proof,
        "raw_price_excel": str(output_excel),
        "row_count": len(rows),
        "mill_count": len(_source_mills(rows)),
        "source_mode": "web",
    }
    report_out.parent.mkdir(parents=True, exist_ok=True)
    report_out.write_text(
        json.dumps(
            report,
            ensure_ascii=False,
            indent=2,
            default=lambda obj: obj.isoformat() if isinstance(obj, (datetime, date)) else str(obj),
        ),
        encoding="utf-8",
    )
    return report

def prepare_web_mapping(
    project_excel: Path,
    raw_excel: Path,
    official_mapping_json: Path,
    pending_mapping_json: Path,
    pending_mapping_csv: Path,
    report_out: Path,
) -> dict[str, Any]:
    source_rows = _load_raw(raw_excel)
    source_mills = _source_mills(source_rows)
    wb_project = load_workbook_safe(project_excel, data_only=True)
    project_sheets = [x for x in wb_project.sheetnames if x not in SKIP_SHEETS]
    existing = _dedupe_mapping_rows(_load_mapping_rows(official_mapping_json)) if official_mapping_json.exists() else []
    existing_by_sheet = {str(r.get("项目文件Sheet") or "").strip(): r for r in existing}

    current_project_set = set(project_sheets)
    current_source_set = {_norm_company(x) for x in source_mills}
    existing_project_set = _mapping_project_set(existing)
    existing_source_set = _mapping_source_set(existing)
    has_pending = any(str(r.get("状态") or "").startswith("待确认") for r in existing)

    same_sets = bool(existing) and existing_project_set == current_project_set and existing_source_set == current_source_set
    reusable = same_sets and not has_pending

    mapping_rows: list[dict[str, str]] = []
    if "报价表" in wb_project.sheetnames:
        mapping_rows.append(
            {
                "项目文件Sheet": "报价表",
                "最新清单厂家Sheet": "",
                "状态": "跳过(汇总页)",
                "说明": "非厂家页",
            }
        )

    for sheet in project_sheets:
        old = existing_by_sheet.get(sheet)
        old_source = str(old.get("最新清单厂家Sheet") or "").strip() if old else ""
        old_status = str(old.get("状态") or "").strip() if old else ""
        old_note = str(old.get("说明") or "").strip() if old else ""
        if old and old_status in {CONFIRMED_WRITE_STATUS, CONFIRMED_SKIP_STATUS}:
            matched_name = _match_source_name(old_source, source_mills)
            if matched_name:
                mapping_rows.append(
                    {
                        "项目文件Sheet": sheet,
                        "最新清单厂家Sheet": matched_name,
                        "状态": old_status,
                        "说明": old_note or "复用已确认映射",
                    }
                )
                continue
        suggested, status, note = _suggest_source_for_sheet(sheet, source_mills)
        mapping_rows.append(
            {
                "项目文件Sheet": sheet,
                "最新清单厂家Sheet": suggested,
                "状态": status,
                "说明": note,
            }
        )

    # 新厂家判定应基于“是否已在本次对照中出现”，而不是“是否已确认”。
    # 否则首次生成待确认时会把已建议映射的厂家误判为“新厂家”。
    used_mapped_sources = _mapping_mapped_sources(mapping_rows)
    new_sources = []
    for mill in source_mills:
        norm = _norm_company(mill)
        if norm not in used_mapped_sources:
            new_sources.append(mill)
            existing_source_row = None
            for row in existing:
                if not str(row.get("项目文件Sheet") or "").strip() and _norm_company(
                    str(row.get("最新清单厂家Sheet") or "").strip()
                ) == norm:
                    existing_source_row = row
                    break
            if existing_source_row:
                mapping_rows.append(existing_source_row)
            else:
                mapping_rows.append(
                    {
                        "项目文件Sheet": "",
                        "最新清单厂家Sheet": mill,
                        "状态": "待确认(新厂家)",
                        "说明": "请确认是否已确认不更新",
                    }
                )

    pending_mapping_json.parent.mkdir(parents=True, exist_ok=True)
    if reusable:
        mapping_rows = existing

    mapping_rows = _dedupe_mapping_rows(mapping_rows)
    pending_count = sum(1 for r in mapping_rows if str(r.get("状态") or "").startswith("待确认"))
    conflict_count = sum(1 for r in mapping_rows if str(r.get("状态") or "").startswith("待确认(冲突)"))
    unmatched_count = sum(1 for r in mapping_rows if str(r.get("状态") or "").startswith("未匹配"))
    needs_confirm = pending_count > 0
    _write_mapping_json_csv(mapping_rows, pending_mapping_json, pending_mapping_csv)

    report = {
        "phase": "web_prepare",
        "project_excel": project_excel.name,
        "raw_price_excel": str(raw_excel),
        "official_mapping_json": str(official_mapping_json),
        "pending_mapping_json": str(pending_mapping_json),
        "pending_mapping_csv": str(pending_mapping_csv),
        "reuse_existing_mapping": reusable,
        "same_manufacturer_sets": same_sets,
        "project_sheet_count": len(project_sheets),
        "source_mill_count": len(source_mills),
        "pending_count": pending_count,
        "conflict_count": conflict_count,
        "unmatched_count": unmatched_count,
        "new_sources": new_sources,
        "needs_user_confirmation": needs_confirm,
    }
    report_out.parent.mkdir(parents=True, exist_ok=True)
    report_out.write_text(
        json.dumps(
            report,
            ensure_ascii=False,
            indent=2,
            default=lambda obj: obj.isoformat() if isinstance(obj, (datetime, date)) else str(obj),
        ),
        encoding="utf-8",
    )
    return report


def _rows_for_source(source_company: str, source_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not source_company:
        return []
    by_mill: dict[str, list[dict[str, Any]]] = {}
    for row in source_rows:
        mill = _clean_place_text(str(row.get("钢厂/产地") or ""))
        if not mill:
            continue
        by_mill.setdefault(mill, []).append(row)

    target_norm = _norm_company(source_company)
    if not target_norm:
        return []
    for mill, rows in by_mill.items():
        if _norm_company(mill) == target_norm:
            return rows

    scored = []
    for mill in by_mill:
        score = _score_company(source_company, mill)
        if score > 0:
            scored.append((score, mill))
    if not scored:
        return []
    scored.sort(key=lambda x: x[0], reverse=True)
    top = [m for sc, m in scored if sc == scored[0][0]]
    if len(top) != 1:
        return []
    return by_mill[top[0]]


def apply_web_writeback(
    project_excel: Path,
    raw_excel: Path,
    mapping_json: Path,
    quote_date: str,
    source_url: str | None,
    report_out: Path,
) -> dict[str, Any]:
    raw_mapping_rows = _load_mapping_rows(mapping_json)
    mapping_rows = _dedupe_mapping_rows(raw_mapping_rows)
    deduped_row_count = max(0, len(raw_mapping_rows) - len(mapping_rows))
    source_rows = _load_raw(raw_excel)
    source_mills = _source_mills(source_rows)
    source_mill_norm = {_norm_company(x) for x in source_mills}
    confirmed_sources = _mapping_confirmed_sources(mapping_rows)

    unresolved_sources = []
    for mill in source_mills:
        if _norm_company(mill) not in confirmed_sources:
            unresolved_sources.append(
                {
                    "来源厂家": mill,
                    "原因": "存在未确认的新厂家，请先更新对照表（可标记为已确认不更新）",
                }
            )

    if unresolved_sources:
        report = {
            "phase": "web_apply",
            "project_excel": project_excel.name,
            "raw_price_excel": str(raw_excel),
            "mapping_json": str(mapping_json),
            "mapping_row_count": len(mapping_rows),
            "deduped_row_count": deduped_row_count,
            "source_quote_date": quote_date,
            "source_url": source_url,
            "blocked": True,
            "blocked_reason": "存在未确认的新厂家",
            "unresolved_sources": unresolved_sources,
            "reset_tab_color": False,
            "unmatched_red_tabs": [],
            "updated_sheet_count": 0,
            "skipped_count": 0,
            "backup_file": None,
            "updated_sheets": [],
            "skipped": [],
        }
        report_out.parent.mkdir(parents=True, exist_ok=True)
        report_out.write_text(
            json.dumps(
                report,
                ensure_ascii=False,
                indent=2,
                default=lambda obj: obj.isoformat() if isinstance(obj, (datetime, date)) else str(obj),
            ),
            encoding="utf-8",
        )
        return report

    wb = load_workbook_safe(project_excel)
    updates: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for sheet in wb.sheetnames:
        wb[sheet].sheet_properties.tabColor = None

    red_tabs: set[str] = set()

    for row in mapping_rows:
        sheet = str(row.get("项目文件Sheet") or "").strip()
        source_company = str(row.get("最新清单厂家Sheet") or "").strip()
        status = str(row.get("状态") or "").strip()
        note = str(row.get("说明") or "").strip()

        if not sheet or sheet in SKIP_SHEETS:
            continue
        if status.startswith("未匹配"):
            red_tabs.add(sheet)
        if status != CONFIRMED_WRITE_STATUS:
            skipped.append(
                {
                    "sheet": sheet,
                    "mill": source_company,
                    "reason": f"状态不是{CONFIRMED_WRITE_STATUS}（当前: {status or '空'}）",
                }
            )
            continue
        if sheet not in wb.sheetnames:
            skipped.append({"sheet": sheet, "mill": source_company, "reason": "项目sheet不存在"})
            continue
        if not source_company:
            skipped.append({"sheet": sheet, "mill": source_company, "reason": "已确认匹配但来源厂家为空"})
            red_tabs.add(sheet)
            continue
        if _norm_company(source_company) not in source_mill_norm:
            maybe = _match_source_name(source_company, source_mills)
            if maybe:
                source_company = maybe
            else:
                skipped.append({"sheet": sheet, "mill": source_company, "reason": "来源厂家在本次清单中不存在"})
                red_tabs.add(sheet)
                continue

        rows = _rows_for_source(source_company, source_rows)
        if not rows:
            skipped.append({"sheet": sheet, "mill": source_company, "reason": "来源厂家在本次清单中无有效行"})
            red_tabs.add(sheet)
            continue
        coil = _pick_coil_10(rows)
        rebar = _pick_rebar_for_mill(rows, sheet)

        ws = wb[sheet]
        old_g1, old_g3, old_g4 = ws["G1"].value, ws["G3"].value, ws["G4"].value
        ws["G1"] = f"网价[{quote_date}]"

        # 只要有一个价格有值就执行回写，缺失的价格保留原值
        has_update = False
        partial_skip = []
        if coil is not None:
            ws["G3"] = coil
            ws["G3"].font = Font(color="FFFF0000")  # 红色标记已更新
            has_update = True
        else:
            partial_skip.append("盘螺")

        if rebar is not None:
            ws["G4"] = rebar
            ws["G4"].font = Font(color="FFFF0000")  # 红色标记已更新
            has_update = True
        else:
            partial_skip.append("螺纹")

        # 如果两个价格都缺失，则跳过
        if not has_update:
            skipped.append({"sheet": sheet, "mill": source_company, "reason": "未提取到盘螺10/螺纹18有效价格"})
            red_tabs.add(sheet)
            continue

        partial_note = note
        if partial_skip:
            partial_note += f"(跳过{'+'.join(partial_skip)}：价格为空，保留原值)"

        updates.append(
            {
                "sheet": sheet,
                "mill": source_company,
                "note": partial_note,
                "G1_old": str(old_g1) if old_g1 is not None else None,
                "G1_new": ws["G1"].value,
                "G3_old": old_g3,
                "G3_coil": coil if coil is not None else f"{old_g3} (保留原值)",
                "G4_old": old_g4,
                "G4_rebar": rebar if rebar is not None else f"{old_g4} (保留原值)",
                "partial_update": bool(partial_skip),
                "skipped_items": partial_skip if partial_skip else None,
            }
        )

    for sheet in sorted(red_tabs):
        if sheet in wb.sheetnames:
            wb[sheet].sheet_properties.tabColor = "FFFF0000"

    backup_dir = project_excel.parent / "备份"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = backup_dir / f"{project_excel.stem}.backup_before_web_write_{stamp}.xlsx"
    shutil.copy2(project_excel, backup_file)
    wb.save(project_excel)

    report = {
        "phase": "web_apply",
        "project_excel": project_excel.name,
        "raw_price_excel": str(raw_excel),
        "mapping_json": str(mapping_json),
        "mapping_row_count": len(mapping_rows),
        "deduped_row_count": deduped_row_count,
        "source_quote_date": quote_date,
        "source_url": source_url,
        "blocked": False,
        "reset_tab_color": True,
        "unmatched_red_tabs": sorted(red_tabs),
        "updated_sheet_count": len(updates),
        "skipped_count": len(skipped),
        "backup_file": str(backup_file),
        "updated_sheets": updates,
        "skipped": skipped,
    }
    report_out.parent.mkdir(parents=True, exist_ok=True)
    report_out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Web price fetch/prepare/apply for project quote workbook.")
    sub = p.add_subparsers(dest="command", required=True)

    p_fetch = sub.add_parser("fetch", help="登录网站并提取网价清单（每厂家一sheet）")
    p_fetch.add_argument("--project", required=True, help="项目报价Excel路径")
    p_fetch.add_argument("--location", help="地点（默认从项目文件名识别）")
    p_fetch.add_argument("--list-url", help="列表页URL（可覆盖默认映射）")
    p_fetch.add_argument("--detail-url", help="详情页URL（当列表页触发校验时可直接指定）")
    p_fetch.add_argument("--account-file", default="网站账号密码.txt", help="账号密码文件")
    p_fetch.add_argument("--username", help="账号")
    p_fetch.add_argument("--password", help="密码")
    p_fetch.add_argument("--output", help="原材料清单输出xlsx路径")
    p_fetch.add_argument("--report-out", help="fetch报告输出路径")
    p_fetch.add_argument("--headless", action="store_true", help="启用无头模式")
    p_fetch.add_argument("--manual-login-timeout", type=int, default=180, help="有头模式下等待人工登录的秒数")

    p_prepare = sub.add_parser("prepare", help="生成网价厂家待确认对照（或复用正式对照）")
    p_prepare.add_argument("--project", required=True, help="项目报价Excel路径")
    p_prepare.add_argument("--raw-excel", required=True, help="网价清单xlsx路径")
    p_prepare.add_argument("--location", help="地点（默认从项目文件名识别）")
    p_prepare.add_argument("--official-json", help="正式对照json路径（默认运行产物/厂家对照表_[地点]_正式.json）")
    p_prepare.add_argument("--pending-json", help="待确认对照json输出路径")
    p_prepare.add_argument("--pending-csv", help="待确认对照csv输出路径")
    p_prepare.add_argument("--report-out", help="prepare报告输出路径")

    p_apply = sub.add_parser("apply", help="按已确认对照回写网价（G1/G3/G4）")
    p_apply.add_argument("--project", required=True, help="项目报价Excel路径")
    p_apply.add_argument("--raw-excel", required=True, help="网价清单xlsx路径")
    p_apply.add_argument("--mapping-json", required=True, help="已确认对照json路径")
    p_apply.add_argument("--quote-date", help="网价日期，格式YYYY-MM-DD")
    p_apply.add_argument("--source-url", help="来源详情页URL")
    p_apply.add_argument("--source-report", help="fetch阶段报告json（用于自动读取quote_date/source_url）")
    p_apply.add_argument("--report-out", help="apply报告输出路径")
    return p


def _load_quote_info(args: argparse.Namespace) -> tuple[str, str | None]:
    quote_date = args.quote_date
    source_url = args.source_url
    if args.source_report:
        payload = json.loads(Path(args.source_report).read_text(encoding="utf-8"))
        if not quote_date:
            quote_date = str(payload.get("quote_date") or "").strip() or None
        if not source_url:
            source_url = str(payload.get("latest_url") or "").strip() or None
    if not quote_date:
        quote_date = datetime.now().strftime("%Y-%m-%d")
    return quote_date, source_url


def main() -> int:
    args = _build_parser().parse_args()

    if args.command == "fetch":
        project = Path(args.project)
        loc = args.location or _parse_location(project.name)[0]
        now_tag = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output = Path(args.output) if args.output else Path("运行产物") / f"_tmp_{loc}_网价清单_{now_tag}.xlsx"
        report_out = (
            Path(args.report_out)
            if args.report_out
            else Path("运行产物") / f"网价提取报告_{loc}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.json"
        )
        report = fetch_web_prices(
            project_excel=project,
            location=args.location,
            list_url=args.list_url,
            detail_url=args.detail_url,
            account_file=Path(args.account_file),
            username=args.username,
            password=args.password,
            output_excel=output,
            report_out=report_out,
            headless=args.headless,
            manual_login_timeout_seconds=args.manual_login_timeout,
        )
        if not args.output:
            target = Path("运行产物") / f"{loc}{report['quote_date']}建筑钢材原料价格清单.xlsx"
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(output), str(target))
            report["raw_price_excel"] = str(target)
            report_out.write_text(
                json.dumps(
                    report,
                    ensure_ascii=False,
                    indent=2,
                    default=lambda obj: obj.isoformat() if isinstance(obj, (datetime, date)) else str(obj),
                ),
                encoding="utf-8",
            )
            output = target
        print(f"Raw Excel: {output}")
        print(f"Quote Date: {report['quote_date']}")
        print(f"Latest URL: {report['latest_url']}")
        print(f"Report: {report_out}")
        return 0

    if args.command == "prepare":
        project = Path(args.project)
        loc = args.location or _parse_location(project.name)[0]
        official_json = (
            Path(args.official_json) if args.official_json else Path("运行产物") / f"厂家对照表_{loc}_正式.json"
        )
        pending_json = (
            Path(args.pending_json) if args.pending_json else Path("运行产物") / f"厂家对照表_{loc}_待确认.json"
        )
        pending_csv = (
            Path(args.pending_csv) if args.pending_csv else Path("运行产物") / f"厂家对照表_{loc}_待确认.csv"
        )
        report_out = (
            Path(args.report_out)
            if args.report_out
            else Path("运行产物") / f"网价回写准备报告_{loc}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.json"
        )
        report = prepare_web_mapping(
            project_excel=project,
            raw_excel=Path(args.raw_excel),
            official_mapping_json=official_json,
            pending_mapping_json=pending_json,
            pending_mapping_csv=pending_csv,
            report_out=report_out,
        )
        print(f"Mapping JSON: {report['pending_mapping_json']}")
        print(f"Mapping CSV: {report['pending_mapping_csv']}")
        print(f"Reuse Existing: {report['reuse_existing_mapping']}")
        print(f"Pending: {report['pending_count']}")
        print(f"Conflicts: {report['conflict_count']}")
        print(f"Needs Confirm: {report['needs_user_confirmation']}")
        print(f"Report: {report_out}")
        return 0

    project = Path(args.project)
    quote_date, source_url = _load_quote_info(args)
    try:
        loc = _parse_location(project.name)[0]
    except WebPriceError:
        loc = "未知地点"
    report_out = (
        Path(args.report_out)
        if args.report_out
        else Path("运行产物") / f"网价回写报告_{loc}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.json"
    )
    report = apply_web_writeback(
        project_excel=project,
        raw_excel=Path(args.raw_excel),
        mapping_json=Path(args.mapping_json),
        quote_date=quote_date,
        source_url=source_url,
        report_out=report_out,
    )
    if report.get("blocked"):
        print("Blocked: 存在未确认的新厂家，未执行写价。")
        for item in report.get("unresolved_sources", []):
            print(f"- {item.get('来源厂家')}: {item.get('原因')}")
    else:
        print(f"Updated: {report['updated_sheet_count']}")
        print(f"Skipped: {report['skipped_count']}")
        print(f"Reset Tab Color: {report['reset_tab_color']}")
        print(f"Unmatched Red Tabs: {', '.join(report['unmatched_red_tabs']) if report['unmatched_red_tabs'] else '(none)'}")
        print(f"Backup: {report['backup_file']}")
    print(f"Report: {report_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
